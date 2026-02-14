from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.http import HttpRequest                                          # IMPORTS FOR HTTP REQUESTS AND RESPONSES
from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt, ensure_csrf_cookie
from django.contrib import messages                                          # IMPORTS FOR DISPLAYING MESSAGES TO THE USER
from django.contrib.auth.decorators import login_required
from django.contrib.auth.hashers import make_password
from django.contrib.auth import authenticate, login, logout
from django.db.models import Q, Sum, Max, Min, F, Count,Avg                  # IMPORTS FOR COMPLEX QUERIES
import json                                                                  # REMEMBER TO REVIEW THIS GUY AT THE ENDPOINTS
import io
import math
from django.db import transaction
from decimal import Decimal
from datetime import datetime,timedelta
import time
from django.db.models.functions import TruncMonth, TruncDate
from django.utils.timezone import now
from django.db.models import F, ExpressionWrapper, DecimalField

from sklearn.ensemble import RandomForestRegressor
import numpy as np
import sklearn
#from sklearn.linear_model import forward_selection
from sklearn.metrics import mean_squared_error, r2_score


# ReportLab imports for PDF generation
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

# openpyxl imports for Excel generation
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

from .models import ItemType,ItemCategory,ItemSubcategory,PaymentMode,County,Town,PaymentStatus,ReceiptStatus,ShippingStatus,UserRole
from .models import Inventory,InventoryItem
from .models import Supplier,PurchaseOrder,PurchaseDetail,Payment
from .models import Customer,SalesOrder,SalesDetail,Receipt
from .models import UserManager,User


# A TEST VIEW TO SEE WHETHER THE Test.html DOCUMENT IS PERFECTLY LOADING
def test_page(request):
    return render(request, 'Test.html')

# ========================================== AUTHENTICATION VIEWS ==============================================================================================
@ensure_csrf_cookie
def login_view(request):
    """
    Handle login - both form display and authentication
    Prevents caching to stop back-button issues
    """
    print("Login view called!")  
    print("Method:", request.method) 
    
    # If user is already authenticated, redirect to index
    if request.user.is_authenticated:
        return redirect('index')
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            email = data.get('email', '').strip().lower()
            password = data.get('password', '').strip()
            
            # INPUT VALIDATION
            if not email or not password:
                return JsonResponse({
                    'success': False,
                    'message': 'Email and password are required'
                }, status=400)
            
            # FIND THE USER USING THE EMAIL
            try:
                user = User.objects.get(email=email)
            except User.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'message': 'Invalid email or password'
                }, status=401)
            
            # CHECK THE PASSWORD
            if not user.check_password(password):
                return JsonResponse({
                    'success': False,
                    'message': 'Invalid email or password'
                }, status=401)
            
            # CONFIRM WHETHER THE USER IS ACTIVE
            if not user.is_active:
                return JsonResponse({
                    'success': False,
                    'message': 'Account is inactive. Please contact administrator.'
                }, status=403)
            
            # LOG-IN THE USER
            login(request, user)
            
            print(f"✅ User {user.full_name} logged in successfully")
            
            return JsonResponse({
                'success': True,
                'message': 'Login successful',
                'username': user.full_name,
                'redirect_url': '/'
            })
            
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'message': 'Invalid request format'
            }, status=400)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'An error occurred: {str(e)}'
            }, status=500)
    
    # GET REQUEST - Render login form with cache control
    response = render(request, 'Log-in_Register.html')
    
    # Prevent browser from caching login page
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    
    return response


def register_view(request):                   # HANDLING BOTH THE USER REGISTRATION FORM AND THE REGISTRATION PROCESS
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # EXTRACT THEN VALIDATE THE DATA
            full_name = data.get('full_name', '').strip()
            email = data.get('email', '').strip().lower()
            phone_number = data.get('phone_number', '').strip()
            password = data.get('password', '').strip()
            role = data.get('role', '').strip()
            
            # INPUT VALIDATION
            if not full_name or len(full_name) < 2:
                return JsonResponse({
                    'success': False,
                    'message': 'Full name must be at least 2 characters'
                }, status=400)
            
            if not email or ('@gmail.com' or '@outlook.com' or '@yahoo.com' or '@students.kcau.ac.ke') not in email:
                return JsonResponse({
                    'success': False,
                    'message': 'Please provide a valid email address'
                }, status=400)
            
            if not phone_number or len(phone_number) < 10:
                return JsonResponse({
                    'success': False,
                    'message': 'Please provide a valid 10-digit phone number'
                }, status=400)
            
            if not password or len(password) < 6:
                return JsonResponse({
                    'success': False,
                    'message': 'Password must be at least 6 characters'
                }, status=400)
            
            if not role:
                return JsonResponse({
                    'success': False,
                    'message': 'Please select a role'
                }, status=400)
            
            # CHECK WHETHER THE E-MAIL EXISTS
            if User.objects.filter(email=email).exists():
                return JsonResponse({
                    'success': False,
                    'message': 'Email already registered'
                }, status=400)
            
            # Get or create UserRole
            user_role, created = UserRole.objects.get_or_create(user_role=role)
            
            # Create user
            user = User.objects.create(
                email=email,
                full_name=full_name,
                phone_number=phone_number,
                user_role=user_role,
                is_active=True
            )
            user.set_password(password)
            user.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Registration successful! Please login.',
                'redirect_url': '/login/'
            })
            
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'message': 'Invalid request format'
            }, status=400)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Registration failed: {str(e)}'
            }, status=500)
    
    return JsonResponse({
        'success': False,
        'message': 'Invalid request method'
    }, status=405)


def logout_view(request):
    """
    Handle user logout
    - Clears Django session
    - Redirects to login page
    """
    # Get username before logout for logging
    username = request.user.full_name if request.user.is_authenticated else 'Unknown'
    
    # Clear Django session
    logout(request)
    
    # Clear any additional session data
    request.session.flush()
    
    print(f"✅ User {username} logged out successfully")
    
    # Redirect to login page
    return redirect('login')

# ================================== MAIN APPLICATION VIEWS ===============================================================================================
"""@login_required(login_url='/login/')"""
def welcome_view(request):
    """Load Welcome page"""
    context = {}
    return render(request, 'Welcome.html', context)

@login_required(login_url='/login/')
def index(request):
    """
    Main view that renders the Index.html with sidebar navigation
    Prevents caching to stop back-button issues after logout
    """
    # Double-check user is authenticated (should be guaranteed by decorator)
    if not request.user.is_authenticated:
        return redirect('login')
    
    context = {
        'username': request.user.full_name
    }
    
    response = render(request, 'Index.html', context)
    
    # Prevent browser from caching this page
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    
    return response

# ================================= DYNAMIC CONTENT VIEWS =================================================================================================

@login_required(login_url='/login/')
def dashboard_content(request):
    """Load Dashboard content"""
    context = {}
    return render(request, 'Dashboard.html', context)


@login_required(login_url='/login/')
def inventory_items_content(request):
    # """Load Inventory Items content"""
    #inventory_items = InventoryItem.objects.all()
    #context = {'inventory_items': inventory_items}
    #return render(request, 'Inventory-items.html', context)
    """Load Inventory Items content - Returns content only, not full page"""
    return render(request, 'Inventory-items.html')

@login_required(login_url='/login/')
def inventory_content(request):
    """Load Inventory content"""
    inventory = Inventory.objects.all()
    context = {'inventory': inventory}
    return render(request, 'Inventory.html', context)


@login_required(login_url='/login/')
def suppliers_content(request):
    """Load Suppliers content"""
    suppliers = Supplier.objects.all()
    context = {'suppliers': suppliers}
    return render(request, 'Suppliers.html', context)


@login_required(login_url='/login/')
def customers_content(request):
    """Load Customers content"""
    customers = Customer.objects.all()
    context = {'customers': customers}
    return render(request, 'Customers.html', context)


@login_required(login_url='/login/')
def purchases_content(request):
    """Load Purchases content"""
    purchase_orders = PurchaseOrder.objects.all()
    context = {'purchase_orders': purchase_orders}
    return render(request, 'Purchases.html', context)


@login_required(login_url='/login/')
def sales_content(request):
    """Load Sales content"""
    sales_orders = SalesOrder.objects.all()
    context = {'sales_orders': sales_orders}
    return render(request, 'Sales.html', context)


@login_required(login_url='/login/')
def receipts_content(request):
    """Load Receipts content"""
    receipts = Receipt.objects.all()
    context = {'receipts': receipts}
    return render(request, 'Receipts.html', context)


@login_required(login_url='/login/')
def payments_content(request):
    """Load Payments content"""
    payments = Payment.objects.all()
    context = {'payments': payments}
    return render(request, 'Payments.html', context)


@login_required(login_url='/login/')
def reports_content(request):
    """Load Reports content"""
    context = {}
    return render(request, 'Reports.html', context)


@login_required(login_url='/login/')
def sales_forecasting_content(request):
    """Load Sales-forecasting content"""
    context = {}
    return render(request, 'Sales-forecasting.html', context)


@login_required(login_url='/login/')
def settings_content(request):
    """Load Settings content"""
    context = {}
    return render(request, 'Settings.html', context)

# ============= API ENDPOINTS =============

@csrf_exempt
@login_required(login_url='/login/')
def api_get_inventory(request):
    """API endpoint to get inventory items"""
    items = Inventory.objects.all().values()
    return JsonResponse(list(items), safe=False)


@csrf_exempt
@login_required(login_url='/login/')
def api_save_inventory(request):
    """API endpoint to save inventory"""
    if request.method == 'POST':
        data = json.loads(request.body)
        # Process and save data
        return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'error'}, status=405)


# =========================== INVENTORY ITEMS VIEWS ======================================================================

@login_required(login_url='/login/')
def inventory_items_content(request):
    """Load Inventory Items content"""
    return render(request, 'Inventory-items.html')


@csrf_exempt
@login_required(login_url='/login/')
def api_get_item_types(request):
    """Get all item types"""
    try:
        types = list(ItemType.objects.values_list('item_type', flat=True))
        return JsonResponse({'success': True, 'data': types})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
@login_required(login_url='/login/')
def api_add_item_type(request):
    """Add new item type"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
    
    try:
        data = json.loads(request.body)
        type_name = data.get('type_name', '').strip()
        
        if not type_name:
            return JsonResponse({'success': False, 'message': 'Item type name is required'}, status=400)
        
        # Check if already exists
        if ItemType.objects.filter(item_type=type_name).exists():
            return JsonResponse({'success': False, 'message': 'Item type already exists'}, status=400)
        
        # Create new item type
        ItemType.objects.create(item_type=type_name)
        
        return JsonResponse({'success': True, 'message': 'Item type added successfully'})
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
@login_required(login_url='/login/')
def api_get_item_categories(request):
    """Get all item categories"""
    try:
        categories = list(ItemCategory.objects.values_list('item_category', flat=True))
        return JsonResponse({'success': True, 'data': categories})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
@login_required(login_url='/login/')
def api_add_item_category(request):
    """Add new item category"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
    
    try:
        data = json.loads(request.body)
        category_name = data.get('category_name', '').strip()
        
        if not category_name:
            return JsonResponse({'success': False, 'message': 'Item category name is required'}, status=400)
        
        # Check if already exists
        if ItemCategory.objects.filter(item_category=category_name).exists():
            return JsonResponse({'success': False, 'message': 'Item category already exists'}, status=400)
        
        # Create new item category
        ItemCategory.objects.create(item_category=category_name)
        
        return JsonResponse({'success': True, 'message': 'Item category added successfully'})
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
@login_required(login_url='/login/')
def api_get_item_subcategories(request):
    """Get all item subcategories"""
    try:
        subcategories = list(ItemSubcategory.objects.values_list('item_subcategory', flat=True))
        return JsonResponse({'success': True, 'data': subcategories})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
@login_required(login_url='/login/')
def api_add_item_subcategory(request):
    """Add new item subcategory"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
    
    try:
        data = json.loads(request.body)
        subcategory_name = data.get('subcategory_name', '').strip()
        
        if not subcategory_name:
            return JsonResponse({'success': False, 'message': 'Item subcategory name is required'}, status=400)
        
        # Check if already exists
        if ItemSubcategory.objects.filter(item_subcategory=subcategory_name).exists():
            return JsonResponse({'success': False, 'message': 'Item subcategory already exists'}, status=400)
        
        # Create new item subcategory
        ItemSubcategory.objects.create(item_subcategory=subcategory_name)
        
        return JsonResponse({'success': True, 'message': 'Item subcategory added successfully'})
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
@login_required(login_url='/login/')
def api_generate_item_id(request):
    """Generate unique item ID in format IT00001"""
    try:
        # Get the maximum item ID
        max_item = InventoryItem.objects.aggregate(Max('item_id'))['item_id__max']
        
        if max_item:
            # Extract numeric part and increment
            num_part = int(max_item[2:]) + 1
        else:
            num_part = 1
        
        # Format as IT00001
        new_id = f"IT{num_part:05d}"
        
        return JsonResponse({'success': True, 'item_id': new_id})
    
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
@login_required(login_url='/login/')
def api_get_inventory_items(request):
    """Get all inventory items"""
    try:
        items = InventoryItem.objects.select_related(
            'item_type', 'item_category', 'item_subcategory'
        ).all()
        
        items_list = []
        for item in items:
            items_list.append({
                'id': item.item_id,
                'type': item.item_type.item_type,
                'category': item.item_category.item_category,
                'subcategory': item.item_subcategory.item_subcategory,
                'name': item.item_name,
                'purchase_price': float(item.purchase_price),
                'sale_price': float(item.sale_price)
            })
        
        return JsonResponse({'success': True, 'data': items_list})
    
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
@login_required(login_url='/login/')
def api_add_inventory_item(request):
    """Add new inventory item"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
    
    try:
        data = json.loads(request.body)
        
        # Extract and validate data
        item_id = data.get('item_id', '').strip()
        item_type_name = data.get('item_type', '').strip()
        item_category_name = data.get('item_category', '').strip()
        item_subcategory_name = data.get('item_subcategory', '').strip()
        item_name = data.get('item_name', '').strip()
        purchase_price = data.get('purchase_price', 0)
        sale_price = data.get('sale_price', 0)
        
        # Validation
        if not all([item_id, item_type_name, item_category_name, item_subcategory_name, item_name]):
            return JsonResponse({
                'success': False, 
                'message': 'All fields are required'
            }, status=400)
        
        # Check if item ID already exists
        if InventoryItem.objects.filter(item_id=item_id).exists():
            return JsonResponse({
                'success': False, 
                'message': 'Item ID already exists'
            }, status=400)
        
        # Get foreign key objects
        try:
            item_type = ItemType.objects.get(item_type=item_type_name)
            item_category = ItemCategory.objects.get(item_category=item_category_name)
            item_subcategory = ItemSubcategory.objects.get(item_subcategory=item_subcategory_name)
        except (ItemType.DoesNotExist, ItemCategory.DoesNotExist, ItemSubcategory.DoesNotExist) as e:
            return JsonResponse({
                'success': False, 
                'message': 'Invalid type, category, or subcategory'
            }, status=400)
        
        # Create inventory item
        with transaction.atomic():
            inventory_item = InventoryItem.objects.create(
                item_id=item_id,
                item_type=item_type,
                item_category=item_category,
                item_subcategory=item_subcategory,
                item_name=item_name,
                purchase_price=purchase_price,
                sale_price=sale_price
            )
            
            # Also create corresponding Inventory record
            Inventory.objects.create(
                item_id=item_id,
                item_type=item_type,
                item_category=item_category,
                item_subcategory=item_subcategory,
                item_name=item_name,
                purchase_price=purchase_price,
                sale_price=sale_price,
                quantity_purchased=0,
                quantity_sold=0,
                reorder_level=0,
                reorder_required='NO'
            )
        
        return JsonResponse({
            'success': True, 
            'message': 'Inventory item added successfully'
        })
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
@login_required(login_url='/login/')
def api_update_inventory_item(request):
    """Update inventory item"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
    
    try:
        data = json.loads(request.body)
        
        item_id = data.get('item_id', '').strip()
        item_type_name = data.get('item_type', '').strip()
        item_category_name = data.get('item_category', '').strip()
        item_subcategory_name = data.get('item_subcategory', '').strip()
        item_name = data.get('item_name', '').strip()
        purchase_price = data.get('purchase_price', 0)
        sale_price = data.get('sale_price', 0)
        
        # Get the item
        try:
            item = InventoryItem.objects.get(item_id=item_id)
        except InventoryItem.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Item not found'}, status=404)
        
        # Get foreign key objects
        try:
            item_type = ItemType.objects.get(item_type=item_type_name)
            item_category = ItemCategory.objects.get(item_category=item_category_name)
            item_subcategory = ItemSubcategory.objects.get(item_subcategory=item_subcategory_name)
        except (ItemType.DoesNotExist, ItemCategory.DoesNotExist, ItemSubcategory.DoesNotExist):
            return JsonResponse({
                'success': False, 
                'message': 'Invalid type, category, or subcategory'
            }, status=400)
        
        # Update item
        with transaction.atomic():
            item.item_type = item_type
            item.item_category = item_category
            item.item_subcategory = item_subcategory
            item.item_name = item_name
            item.purchase_price = purchase_price
            item.sale_price = sale_price
            item.save()
            
            # Also update Inventory record
            try:
                inventory = Inventory.objects.get(item_id=item_id)
                inventory.item_type = item_type
                inventory.item_category = item_category
                inventory.item_subcategory = item_subcategory
                inventory.item_name = item_name
                inventory.purchase_price = purchase_price
                inventory.sale_price = sale_price
                inventory.save()
            except Inventory.DoesNotExist:
                pass
        
        return JsonResponse({
            'success': True, 
            'message': 'Inventory item updated successfully'
        })
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

@csrf_exempt
@login_required(login_url='/login/')
def api_delete_inventory_item(request):
    """
    Delete inventory item only if no stock exists
    Deletes from BOTH Inventory and InventoryItem tables
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
    
    try:
        data = json.loads(request.body)
        item_id = data.get('item_id', '').strip()
        
        if not item_id:
            return JsonResponse({'success': False, 'message': 'Item ID is required'}, status=400)
        
        try:
            # Try to get from Inventory table first
            inventory_item = Inventory.objects.get(item_id=item_id)
            
            # Calculate remaining quantity
            remaining_qty = inventory_item.quantity_purchased - inventory_item.quantity_sold
            
            # ✅ CRITICAL CHECK: Prevent deletion if stock exists
            if remaining_qty > 0:
                return JsonResponse({
                    'success': False,
                    'message': f'❌ Cannot delete item with existing stock!\n\nItem: {inventory_item.item_name}\nRemaining Quantity: {remaining_qty}\n\nPlease sell or adjust the inventory first.',
                    'has_stock': True,
                    'remaining_qty': remaining_qty
                }, status=400)
            
            # ✅ CRITICAL CHECK: Prevent deletion if there were any transactions
            if inventory_item.quantity_purchased > 0 or inventory_item.quantity_sold > 0:
                return JsonResponse({
                    'success': False,
                    'message': f'❌ Cannot delete item with transaction history!\n\nItem: {inventory_item.item_name}\nPurchased: {inventory_item.quantity_purchased}\nSold: {inventory_item.quantity_sold}\n\nThis item has existing transactions.',
                    'has_transactions': True
                }, status=400)
            
            # ✅ Safe to delete - no stock and no transactions
            with transaction.atomic():
                # Delete from InventoryItem table
                try:
                    inventory_master = InventoryItem.objects.get(item_id=item_id)
                    inventory_master.delete()
                    print(f"✅ Deleted from InventoryItem: {item_id}")
                except InventoryItem.DoesNotExist:
                    print(f"⚠️ Item {item_id} not found in InventoryItem table (already deleted or never existed)")
                
                # Delete from Inventory table
                inventory_item.delete()
                print(f"✅ Deleted from Inventory: {item_id}")
            
            return JsonResponse({
                'success': True,
                'message': f'✅ Item deleted successfully!\n\nItem ID: {item_id}\nItem Name: {inventory_item.item_name}'
            })
        
        except Inventory.DoesNotExist:
            # Item not found in Inventory, check InventoryItem
            try:
                inventory_master = InventoryItem.objects.get(item_id=item_id)
                
                # If found in InventoryItem only, safe to delete (no transactions ever happened)
                with transaction.atomic():
                    inventory_master.delete()
                    print(f"✅ Deleted from InventoryItem only: {item_id}")
                
                return JsonResponse({
                    'success': True,
                    'message': f'✅ Item deleted successfully!\n\nItem ID: {item_id}'
                })
            
            except InventoryItem.DoesNotExist:
                return JsonResponse({
                    'success': False, 
                    'message': f'❌ Item not found!\n\nItem ID: {item_id}'
                }, status=404)
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        print(f"❌ Error deleting inventory item: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'message': f'Server error: {str(e)}'}, status=500)
    
# ================================= ADDED FOR THE INVENTORY MODULE ============================================
    # Get all inventory with calculated fields
@csrf_exempt
@login_required(login_url='/login/')
def api_get_inventory(request):
    """Get all inventory items with quantities and reorder status"""
    try:
        inventory_items = Inventory.objects.select_related(
            'item_type', 'item_category', 'item_subcategory'
        ).all()
        
        items_list = []
        for item in inventory_items:
            remaining_qty = item.quantity_purchased - item.quantity_sold
            reorder_required = 'YES' if remaining_qty <= item.reorder_level else 'NO'
            
            items_list.append({
                'id': item.item_id,
                'type': item.item_type.item_type,
                'category': item.item_category.item_category,
                'subcategory': item.item_subcategory.item_subcategory,
                'name': item.item_name,
                'purchasedQty': item.quantity_purchased,
                'soldQty': item.quantity_sold,
                'remainingQty': remaining_qty,
                'reorderLevel': item.reorder_level,
                'reorderRequired': reorder_required,
                'purchase_price': float(item.purchase_price),  
                'sale_price': float(item.sale_price)           
            })
        
        return JsonResponse({'success': True, 'data': items_list})
    
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


# Update reorder level only
@csrf_exempt
@login_required(login_url='/login/')
def api_update_reorder_level(request):
    """Update only the reorder level for an inventory item"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
    
    try:
        data = json.loads(request.body)
        item_id = data.get('item_id', '').strip()
        reorder_level = int(data.get('reorder_level', 0))
        
        if not item_id:
            return JsonResponse({'success': False, 'message': 'Item ID is required'}, status=400)
        
        if reorder_level < 0:
            return JsonResponse({'success': False, 'message': 'Reorder level must be non-negative'}, status=400)
        
        try:
            item = Inventory.objects.get(item_id=item_id)
            item.reorder_level = reorder_level
            item.save()  # This will trigger the model's save() method to update reorder_required
            
            return JsonResponse({
                'success': True,
                'message': 'Reorder level updated successfully'
            })
        
        except Inventory.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Item not found'}, status=404)
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except ValueError:
        return JsonResponse({'success': False, 'message': 'Reorder level must be a number'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


# Delete inventory item (with validation)
@csrf_exempt
@login_required(login_url='/login/')
def api_delete_inventory_item(request):
    """Delete inventory item only if no stock exists"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
    
    try:
        data = json.loads(request.body)
        item_id = data.get('item_id', '').strip()
        
        if not item_id:
            return JsonResponse({'success': False, 'message': 'Item ID is required'}, status=400)
        
        try:
            item = Inventory.objects.get(item_id=item_id)
            
            # Check if item has stock
            if item.quantity_purchased > 0 or item.quantity_sold > 0:
                return JsonResponse({
                    'success': False,
                    'message': 'Cannot delete item with existing stock transactions'
                }, status=400)
            
            # Delete from both tables
            with transaction.atomic():
                InventoryItem.objects.filter(item_id=item_id).delete()
                item.delete()
            
            return JsonResponse({
                'success': True,
                'message': 'Inventory item deleted successfully'
            })
        
        except Inventory.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Item not found'}, status=404)
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)
    

# ===================== SUPPLIERS MODULE VIEWS =====================

@login_required(login_url='/login/')
def suppliers_content(request):
    """Load Suppliers content page"""
    return render(request, 'Suppliers.html')


# ===================== API ENDPOINTS =====================

@csrf_exempt
@login_required(login_url='/login/')
def api_get_counties(request):
    """Get all counties from database"""
    try:
        counties = list(County.objects.values_list('county', flat=True).order_by('county'))
        return JsonResponse({'success': True, 'data': counties})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
@login_required(login_url='/login/')
def api_get_towns(request):
    """Get all towns from database"""
    try:
        towns = list(Town.objects.values_list('town', flat=True).order_by('town'))
        return JsonResponse({'success': True, 'data': towns})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
@login_required(login_url='/login/')
def api_add_county(request):
    """Add new county"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
    
    try:
        data = json.loads(request.body)
        county_name = data.get('county_name', '').strip()
        
        if not county_name:
            return JsonResponse({'success': False, 'message': 'County name is required'}, status=400)
        
        # Check if already exists
        if County.objects.filter(county__iexact=county_name).exists():
            return JsonResponse({'success': False, 'message': 'County already exists'}, status=400)
        
        # Create new county
        County.objects.create(county=county_name)
        
        return JsonResponse({'success': True, 'message': 'County added successfully'})
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
@login_required(login_url='/login/')
def api_add_town(request):
    """Add new town"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
    
    try:
        data = json.loads(request.body)
        town_name = data.get('town_name', '').strip()
        
        if not town_name:
            return JsonResponse({'success': False, 'message': 'Town name is required'}, status=400)
        
        # Check if already exists
        if Town.objects.filter(town__iexact=town_name).exists():
            return JsonResponse({'success': False, 'message': 'Town already exists'}, status=400)
        
        # Create new town
        Town.objects.create(town=town_name)
        
        return JsonResponse({'success': True, 'message': 'Town added successfully'})
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
@login_required(login_url='/login/')
def api_generate_supplier_id(request):
    """Generate unique supplier ID in format S00001"""
    try:
        # Get the maximum supplier ID
        max_supplier = Supplier.objects.aggregate(Max('supplier_id'))['supplier_id__max']
        
        if max_supplier:
            # Extract numeric part and increment
            num_part = int(max_supplier[5:]) + 1
        else:
            num_part = 1
        
        # Format as SUP001
        new_id = f"S{num_part:05d}"
        
        # Make sure it doesn't exist (safety check)
        while Supplier.objects.filter(supplier_id=new_id).exists():
            num_part += 1
            new_id = f"S{num_part:05d}"
        
        return JsonResponse({'success': True, 'supplier_id': new_id})
    
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
@login_required(login_url='/login/')
def api_get_suppliers(request):
    """Get all suppliers with calculated totals"""
    try:
        suppliers = Supplier.objects.select_related('county', 'town').all()
        
        suppliers_list = []
        for supplier in suppliers:
            # Calculate totals from PurchaseOrders
            purchase_totals = PurchaseOrder.objects.filter(
                supplier_id=supplier
            ).aggregate(
                total_purchases=Sum('total_amount'),
                total_payments=Sum('amount_paid')
            )
            
            total_purchases = purchase_totals['total_purchases'] or Decimal('0.00')
            total_payments = purchase_totals['total_payments'] or Decimal('0.00')
            balance = total_purchases - total_payments
            
            # Update supplier record with calculated values
            supplier.total_purchases = total_purchases
            supplier.total_payments = total_payments
            supplier.save()
            
            suppliers_list.append({
                'id': supplier.supplier_id,
                'name': supplier.supplier_name,
                'contact': supplier.phone_number or '',
                'email': supplier.email or '',
                'state': supplier.county.county if supplier.county else '',
                'city': supplier.town.town if supplier.town else '',
                'purchases': float(total_purchases),
                'payments': float(total_payments),
                'balance': float(balance)
            })
        
        return JsonResponse({'success': True, 'data': suppliers_list})
    
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
@login_required(login_url='/login/')
def api_add_supplier(request):
    """Add new supplier"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
    
    try:
        data = json.loads(request.body)
        
        # Extract and validate data
        supplier_id = data.get('id', '').strip()
        supplier_name = data.get('name', '').strip()
        phone_number = data.get('contact', '').strip()
        email = data.get('email', '').strip()
        county_name = data.get('state', '').strip()
        town_name = data.get('city', '').strip()
        
        # Validation
        if not all([supplier_id, supplier_name, county_name, town_name]):
            return JsonResponse({
                'success': False, 
                'message': 'Supplier ID, Name, County, and Town are required'
            }, status=400)
        
        # Check if supplier ID already exists
        if Supplier.objects.filter(supplier_id=supplier_id).exists():
            return JsonResponse({
                'success': False, 
                'message': 'Supplier ID already exists'
            }, status=400)
        
        # Get foreign key objects
        try:
            county = County.objects.get(county=county_name)
            town = Town.objects.get(town=town_name)
        except (County.DoesNotExist, Town.DoesNotExist):
            return JsonResponse({
                'success': False, 
                'message': 'Invalid county or town'
            }, status=400)
        
        # Create supplier
        Supplier.objects.create(
            supplier_id=supplier_id,
            supplier_name=supplier_name,
            phone_number=phone_number if phone_number else None,
            email=email if email else None,
            county=county,
            town=town,
            total_purchases=Decimal('0.00'),
            total_payments=Decimal('0.00')
        )
        
        return JsonResponse({
            'success': True, 
            'message': 'Supplier added successfully'
        })
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
@login_required(login_url='/login/')
def api_update_supplier(request):
    """Update supplier information"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
    
    try:
        data = json.loads(request.body)
        
        supplier_id = data.get('id', '').strip()
        supplier_name = data.get('name', '').strip()
        phone_number = data.get('contact', '').strip()
        email = data.get('email', '').strip()
        county_name = data.get('state', '').strip()
        town_name = data.get('city', '').strip()
        
        # Get the supplier
        try:
            supplier = Supplier.objects.get(supplier_id=supplier_id)
        except Supplier.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Supplier not found'}, status=404)
        
        # Get foreign key objects
        try:
            county = County.objects.get(county=county_name)
            town = Town.objects.get(town=town_name)
        except (County.DoesNotExist, Town.DoesNotExist):
            return JsonResponse({
                'success': False, 
                'message': 'Invalid county or town'
            }, status=400)
        
        # Update supplier
        supplier.supplier_name = supplier_name
        supplier.phone_number = phone_number if phone_number else None
        supplier.email = email if email else None
        supplier.county = county
        supplier.town = town
        supplier.save()
        
        return JsonResponse({
            'success': True, 
            'message': 'Supplier updated successfully'
        })
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
@login_required(login_url='/login/')
def api_delete_supplier(request):
    """Delete supplier (only if balance is zero)"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
    
    try:
        data = json.loads(request.body)
        supplier_id = data.get('supplier_id', '').strip()
        
        if not supplier_id:
            return JsonResponse({'success': False, 'message': 'Supplier ID is required'}, status=400)
        
        try:
            supplier = Supplier.objects.get(supplier_id=supplier_id)
            
            # Calculate balance
            purchase_totals = PurchaseOrder.objects.filter(
                supplier_id=supplier
            ).aggregate(
                total_purchases=Sum('total_amount'),
                total_payments=Sum('amount_paid')
            )
            
            total_purchases = purchase_totals['total_purchases'] or Decimal('0.00')
            total_payments = purchase_totals['total_payments'] or Decimal('0.00')
            balance = total_purchases - total_payments
            
            # Check if balance is zero
            if balance > Decimal('0.00'):
                return JsonResponse({
                    'success': False, 
                    'message': 'Cannot delete supplier with outstanding balance. Please clear all dues first.',
                    'has_balance': True
                }, status=400)
            
            # Delete supplier
            supplier.delete()
            
            return JsonResponse({
                'success': True, 
                'message': 'Supplier deleted successfully'
            })
        
        except Supplier.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Supplier not found'}, status=404)
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
    
# ========================================== CUSTOMER MODULE VIEWS ================================================================================

@login_required(login_url='/login/')
def customers_content(request):
    """Load Customers content page"""
    return render(request, 'Customers.html')


# ================================================ API ENDPOINTS ====================================================================================

@csrf_exempt
@login_required(login_url='/login/')
def api_get_counties(request):
    """Get all counties from database"""
    try:
        counties = list(County.objects.values_list('county', flat=True).order_by('county'))
        return JsonResponse({'success': True, 'data': counties})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
@login_required(login_url='/login/')
def api_get_towns(request):
    """Get all towns from database"""
    try:
        towns = list(Town.objects.values_list('town', flat=True).order_by('town'))
        return JsonResponse({'success': True, 'data': towns})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
@login_required(login_url='/login/')
def api_add_county(request):
    """Add new county"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
    
    try:
        data = json.loads(request.body)
        county_name = data.get('county_name', '').strip()
        
        if not county_name:
            return JsonResponse({'success': False, 'message': 'County name is required'}, status=400)
        
        # Check if already exists
        if County.objects.filter(county__iexact=county_name).exists():
            return JsonResponse({'success': False, 'message': 'County already exists'}, status=400)
        
        # Create new county
        County.objects.create(county=county_name)
        
        return JsonResponse({'success': True, 'message': 'County added successfully'})
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
@login_required(login_url='/login/')
def api_add_town(request):
    """Add new town"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
    
    try:
        data = json.loads(request.body)
        town_name = data.get('town_name', '').strip()
        
        if not town_name:
            return JsonResponse({'success': False, 'message': 'Town name is required'}, status=400)
        
        # Check if already exists
        if Town.objects.filter(town__iexact=town_name).exists():
            return JsonResponse({'success': False, 'message': 'Town already exists'}, status=400)
        
        # Create new town
        Town.objects.create(town=town_name)
        
        return JsonResponse({'success': True, 'message': 'Town added successfully'})
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
@login_required(login_url='/login/')
def api_generate_customer_id(request):
    """Generate unique customer ID in format C00001"""
    try:
        # Get the maximum customer ID
        max_customer = Customer.objects.aggregate(Max('customer_id'))['customer_id__max']
        
        if max_customer:
            # Extract numeric part and increment
            # Handle both C00001 and CUST00001 formats
            if max_customer.startswith('C'):
                num_part = int(max_customer[4:]) + 1
            else:
                num_part = int(max_customer[1:]) + 1
        else:
            num_part = 1
        
        # Format as C00001
        new_id = f"C{num_part:05d}"
        
        # Make sure it doesn't exist (safety check)
        while Customer.objects.filter(customer_id=new_id).exists():
            num_part += 1
            new_id = f"C{num_part:05d}"
        
        return JsonResponse({'success': True, 'customer_id': new_id})
    
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
@login_required(login_url='/login/')
def api_get_customers(request):
    """
    Get all customers with calculated totals
    
    Logic:
    - total_sales = Sum of all SalesOrder.total_amount for this customer
    - total_payments = Sum of all Receipt.amount_received for this customer
    - balance_receivable = total_sales - total_payments
    """
    try:
        customers = Customer.objects.select_related('county', 'town').all()
        
        customers_list = []
        for customer in customers:
            # Calculate total sales from SalesOrders
            sales_totals = SalesOrder.objects.filter(
                customer_id=customer
            ).aggregate(
                total_sales=Sum('total_amount')
            )
            
            # Calculate total payments from Receipts
            payment_totals = Receipt.objects.filter(
                customer_id=customer
            ).aggregate(
                total_payments=Sum('amount_received')
            )
            
            total_sales = sales_totals['total_sales'] or Decimal('0.00')
            total_payments = payment_totals['total_payments'] or Decimal('0.00')
            balance_receivable = total_sales - total_payments
            
            # Update customer record with calculated values
            customer.total_sales = total_sales
            customer.total_payments = total_payments
            customer.save()
            
            customers_list.append({
                'id': customer.customer_id,
                'name': customer.customer_name,
                'contact': customer.phone_number or '',
                'email': customer.email or '',
                'state': customer.county.county if customer.county else '',
                'city': customer.town.town if customer.town else '',
                'sales': float(total_sales),
                'receipts': float(total_payments),  # This is what we send to frontend
                'balance': float(balance_receivable)
            })
        
        return JsonResponse({'success': True, 'data': customers_list})
    
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
@login_required(login_url='/login/')
def api_add_customer(request):
    """Add new customer"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
    
    try:
        data = json.loads(request.body)
        
        # Extract and validate data
        customer_id = data.get('id', '').strip()
        customer_name = data.get('name', '').strip()
        phone_number = data.get('contact', '').strip()
        email = data.get('email', '').strip()
        county_name = data.get('state', '').strip()
        town_name = data.get('city', '').strip()
        
        # Validation
        if not all([customer_id, customer_name, county_name, town_name]):
            return JsonResponse({
                'success': False, 
                'message': 'Customer ID, Name, County, and Town are required'
            }, status=400)
        
        # Check if customer ID already exists
        if Customer.objects.filter(customer_id=customer_id).exists():
            return JsonResponse({
                'success': False, 
                'message': 'Customer ID already exists'
            }, status=400)
        
        # Get foreign key objects
        try:
            county = County.objects.get(county=county_name)
            town = Town.objects.get(town=town_name)
        except (County.DoesNotExist, Town.DoesNotExist):
            return JsonResponse({
                'success': False, 
                'message': 'Invalid county or town'
            }, status=400)
        
        # Create customer with correct field names
        Customer.objects.create(
            customer_id=customer_id,
            customer_name=customer_name,
            phone_number=phone_number if phone_number else None,
            email=email if email else None,
            county=county,
            town=town,
            total_sales=Decimal('0.00'),
            total_payments=Decimal('0.00')  # CORRECTED: Using total_payments not total_receipts
        )
        
        return JsonResponse({
            'success': True, 
            'message': 'Customer added successfully'
        })
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
@login_required(login_url='/login/')
def api_update_customer(request):
    """Update customer information"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
    
    try:
        data = json.loads(request.body)
        
        customer_id = data.get('id', '').strip()
        customer_name = data.get('name', '').strip()
        phone_number = data.get('contact', '').strip()
        email = data.get('email', '').strip()
        county_name = data.get('state', '').strip()
        town_name = data.get('city', '').strip()
        
        # Get the customer
        try:
            customer = Customer.objects.get(customer_id=customer_id)
        except Customer.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Customer not found'}, status=404)
        
        # Get foreign key objects
        try:
            county = County.objects.get(county=county_name)
            town = Town.objects.get(town=town_name)
        except (County.DoesNotExist, Town.DoesNotExist):
            return JsonResponse({
                'success': False, 
                'message': 'Invalid county or town'
            }, status=400)
        
        # Update customer
        customer.customer_name = customer_name
        customer.phone_number = phone_number if phone_number else None
        customer.email = email if email else None
        customer.county = county
        customer.town = town
        customer.save()
        
        return JsonResponse({
            'success': True, 
            'message': 'Customer updated successfully'
        })
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
@login_required(login_url='/login/')
def api_delete_customer(request):
    """
    Delete customer (only if balance is zero)
    
    Logic:
    - Calculate total_sales from all SalesOrders
    - Calculate total_payments from all Receipts
    - If balance (total_sales - total_payments) > 0, prevent deletion
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
    
    try:
        data = json.loads(request.body)
        customer_id = data.get('customer_id', '').strip()
        
        if not customer_id:
            return JsonResponse({'success': False, 'message': 'Customer ID is required'}, status=400)
        
        try:
            customer = Customer.objects.get(customer_id=customer_id)
            
            # Calculate balance from SalesOrders and Receipts
            sales_totals = SalesOrder.objects.filter(
                customer_id=customer
            ).aggregate(
                total_sales=Sum('total_amount')
            )
            
            payment_totals = Receipt.objects.filter(
                customer_id=customer
            ).aggregate(
                total_payments=Sum('amount_received')
            )
            
            total_sales = sales_totals['total_sales'] or Decimal('0.00')
            total_payments = payment_totals['total_payments'] or Decimal('0.00')
            balance_receivable = total_sales - total_payments
            
            # Check if balance is zero
            if balance_receivable > Decimal('0.00'):
                return JsonResponse({
                    'success': False, 
                    'message': f'Customer has outstanding balance of {float(balance_receivable):.2f}. Please clear all dues first.',
                    'has_balance': True,
                    'balance_amount': float(balance_receivable)
                }, status=400)
            
            # Delete customer
            customer.delete()
            
            return JsonResponse({
                'success': True, 
                'message': 'Customer deleted successfully'
            })
        
        except Customer.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Customer not found'}, status=404)
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
    
# ============================================ PURCHASES MODULE VIEWS ======================================================================================

@login_required(login_url='/login/')
def purchases_content(request):
    """Load Purchases content page"""
    return render(request, 'Purchases.html')

# ================================================ PAYMENT STATUS APIs =====================================================================================

@csrf_exempt
@login_required(login_url='/login/')
def api_get_payment_statuses(request):
    """Get all payment statuses"""
    try:
        statuses = list(PaymentStatus.objects.values_list('payment_status', flat=True).order_by('payment_status'))
        return JsonResponse({'success': True, 'data': statuses})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

@csrf_exempt
@login_required(login_url='/login/')
def api_add_payment_status(request):
    """Add new payment status"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
    
    try:
        data = json.loads(request.body)
        status_name = data.get('status_name', '').strip()
        
        if not status_name:
            return JsonResponse({'success': False, 'message': 'Payment status name is required'}, status=400)
        
        if PaymentStatus.objects.filter(payment_status__iexact=status_name).exists():
            return JsonResponse({'success': False, 'message': 'Payment status already exists'}, status=400)
        
        PaymentStatus.objects.create(payment_status=status_name)
        
        return JsonResponse({'success': True, 'message': 'Payment status added successfully'})
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

# ================================================== SHIPPING STATUS APIs =======================================================================================

@csrf_exempt
@login_required(login_url='/login/')
def api_get_shipping_statuses(request):
    """Get all shipping statuses"""
    try:
        statuses = list(ShippingStatus.objects.values_list('shipping_status', flat=True).order_by('shipping_status'))
        return JsonResponse({'success': True, 'data': statuses})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

@csrf_exempt
@login_required(login_url='/login/')
def api_add_shipping_status(request):
    """Add new shipping status"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
    
    try:
        data = json.loads(request.body)
        status_name = data.get('status_name', '').strip()
        
        if not status_name:
            return JsonResponse({'success': False, 'message': 'Shipping status name is required'}, status=400)
        
        if ShippingStatus.objects.filter(shipping_status__iexact=status_name).exists():
            return JsonResponse({'success': False, 'message': 'Shipping status already exists'}, status=400)
        
        ShippingStatus.objects.create(shipping_status=status_name)
        
        return JsonResponse({'success': True, 'message': 'Shipping status added successfully'})
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


# =================================== PURCHASE ORDER APIs ===================================================================================================

# ================================ GENERATING PURCHASE ORDER ID =============================================================================================
@csrf_exempt
@login_required(login_url='/login/')
def api_generate_po_id(request):
    """Generate unique PO ID in format PO00001 and corresponding Bill Number B00001"""
    try:
        max_po = PurchaseOrder.objects.aggregate(Max('po_id'))['po_id__max']
        
        if max_po:
            num_part = int(max_po[2:]) + 1
        else:
            num_part = 1
        
        new_po_id = f"PO{num_part:05d}"
        
        # Make sure it doesn't exist
        while PurchaseOrder.objects.filter(po_id=new_po_id).exists():
            num_part += 1
            new_po_id = f"PO{num_part:05d}"
        
        # Generate corresponding bill number
        bill_number = f"B{num_part:05d}"
        
        return JsonResponse({
            'success': True, 
            'po_id': new_po_id,
            'bill_number': bill_number
        })
    
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

# ========================== GENERATING DETAIL ID FOR THE PURCHASE DETAILS ============================================================================================
@csrf_exempt
@login_required(login_url='/login/')
def api_generate_detail_id(request):
    """
    Generate unique Detail ID for PURCHASES in format PD00001
    Each call generates a FRESH unique ID
    """
    try:
        # Always fetch fresh from database
        max_purchase_detail = PurchaseDetail.objects.aggregate(Max('detail_id'))['detail_id__max']
        
        if max_purchase_detail:
            # Extract numeric part (skip 'PD' prefix)
            num_part = int(max_purchase_detail[2:]) + 1
        else:
            # First purchase detail ever
            num_part = 1
        
        # Format as PD00001
        new_detail_id = f"PD{num_part:05d}"
        
        # Safety check: ensure it doesn't exist
        attempts = 0
        while PurchaseDetail.objects.filter(detail_id=new_detail_id).exists():
            num_part += 1
            new_detail_id = f"PD{num_part:05d}"
            attempts += 1
            if attempts > 1000:
                return JsonResponse({'success': False, 'message': 'Could not generate unique Detail ID'}, status=500)
        
        print(f"✅ GENERATED PURCHASE DETAIL ID: {new_detail_id} at {time.strftime('%H:%M:%S')}")
        
        return JsonResponse({'success': True, 'detail_id': new_detail_id})
    
    except Exception as e:
        print(f"❌ ERROR generating Detail ID: {str(e)}")
        return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
# ================= GET THE NEXT DETAIL NUMBER-FOR NUMBER CONTEXTUALIZATION, NOT THE ACTUAL DETAIL ID ==============================================================
@csrf_exempt
@login_required(login_url='/login/')
def api_get_next_detail_number(request):
    """
    Get the next available detail number for BOTH Purchases and Sales
    This returns just the number, not the full ID
    """
    try:
        max_purchase_detail = PurchaseDetail.objects.aggregate(Max('detail_id'))['detail_id__max']
        max_sales_detail = SalesDetail.objects.aggregate(Max('detail_id'))['detail_id__max']
        
        # Compare both and get the highest
        max_details = []
        if max_purchase_detail:
            max_details.append(int(max_purchase_detail[2:]))  # Skip 'PD' prefix
        if max_sales_detail:
            max_details.append(int(max_sales_detail[2:]))  # Skip 'SD' prefix
        
        if max_details:
            next_number = max(max_details) + 1
        else:
            next_number = 1
        
        return JsonResponse({'success': True, 'next_number': next_number})
    
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
    
@csrf_exempt
@login_required(login_url='/login/')
def api_get_purchase_orders(request):
    """Get all purchase orders with calculated totals"""
    try:
        purchase_orders = PurchaseOrder.objects.select_related(
            'supplier_id', 'county', 'town', 'payment_status', 'shipping_status'
        ).all().order_by('-date')
        
        po_list = []
        for po in purchase_orders:
            # Format date as DD/MM/YYYY
            date_str = po.date.strftime('%d/%m/%Y') if po.date else ''
            
            po_list.append({
                'po_id': po.po_id,
                'date': date_str,
                'supplier_id': po.supplier_id.supplier_id if po.supplier_id else '',
                'supplier_name': po.supplier_name,
                'bill_number': po.bill_number,
                'county': po.county.county if po.county else '',
                'town': po.town.town if po.town else '',
                'total_amount': float(po.total_amount),
                'amount_paid': float(po.amount_paid),
                'balance_left': float(po.total_amount - po.amount_paid),
                'payment_status': po.payment_status.payment_status if po.payment_status else '',
                'shipping_status': po.shipping_status.shipping_status if po.shipping_status else ''
            })
        
        return JsonResponse({'success': True, 'data': po_list})
    
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

@csrf_exempt
@login_required(login_url='/login/')
def api_get_po_details(request, po_id):
    """Get purchase order details including all line items"""
    try:
        # Get purchase order
        po = PurchaseOrder.objects.select_related(
            'supplier_id', 'county', 'town', 'payment_status', 'shipping_status'
        ).get(po_id=po_id)
        
        # Format date
        date_str = po.date.strftime('%d/%m/%Y') if po.date else ''
        
        po_data = {
            'po_id': po.po_id,
            'date': date_str,
            'supplier_id': po.supplier_id.supplier_id if po.supplier_id else '',
            'supplier_name': po.supplier_name,
            'bill_number': po.bill_number,
            'county': po.county.county if po.county else '',
            'town': po.town.town if po.town else '',
            'total_amount': float(po.total_amount),
            'amount_paid': float(po.amount_paid),
            'payment_status': po.payment_status.payment_status if po.payment_status else '',
            'shipping_status': po.shipping_status.shipping_status if po.shipping_status else ''
        }
        
        # Get purchase details
        details = PurchaseDetail.objects.filter(po_id=po).order_by('detail_id')
        details_list = []
        
        for detail in details:
            detail_date_str = detail.date.strftime('%d/%m/%Y') if detail.date else ''
            
            details_list.append({
                'detail_id': detail.detail_id,
                'date': detail_date_str,
                'po_id': detail.po_id.po_id,
                'supplier_id': detail.supplier_id.supplier_id if detail.supplier_id else '',
                'supplier_name': detail.supplier_name,
                'county': detail.county.county if detail.county else '',
                'town': detail.town.town if detail.town else '',
                'bill_number': detail.bill_number,
                'item_id': detail.item_id.item_id if detail.item_id else '',
                'item_type': detail.item_type,
                'item_category': detail.item_category,
                'item_subcategory': detail.item_subcategory,
                'item_name': detail.item_name,
                'quantity_purchased': detail.quantity_purchased,
                'unit_cost': float(detail.unit_cost),
                'tax_rate': float(detail.tax_rate),
                'cost_excluding_tax': float(detail.cost_excluding_tax),
                'total_tax': float(detail.total_tax),
                'cost_including_tax': float(detail.cost_including_tax),
                'shipping_fees': float(detail.shipping_fees),
                'total_purchase_price': float(detail.total_purchase_price)
            })
        
        return JsonResponse({
            'success': True,
            'data': {
                'purchase_order': po_data,
                'purchase_details': details_list
            }
        })
    
    except PurchaseOrder.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Purchase order not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

def parse_date(date_str):
    """Parse date from DD/MM/YYYY format"""
    try:
        return datetime.strptime(date_str, '%d/%m/%Y').date()
    except:
        return None

@csrf_exempt
@login_required(login_url='/login/')
def api_add_purchase_order(request):
    """Add new purchase order with automatic status calculation"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
    
    try:
        data = json.loads(request.body)
        
        # Extract PO header data
        po_id = data.get('po_id', '').strip()
        date_str = data.get('date', '').strip()
        supplier_id = data.get('supplier_id', '').strip()
        supplier_name = data.get('supplier_name', '').strip()
        county_name = data.get('county', '').strip()
        town_name = data.get('town', '').strip()
        bill_number = data.get('bill_number', '').strip()
        items = data.get('items', [])
        
        # Validation
        if not all([po_id, date_str, supplier_name, bill_number]):
            return JsonResponse({
                'success': False, 
                'message': 'All PO header fields are required'
            }, status=400)
        
        if not items:
            return JsonResponse({
                'success': False, 
                'message': 'At least one item is required'
            }, status=400)
        
        # Parse date
        po_date = parse_date(date_str)
        if not po_date:
            return JsonResponse({
                'success': False, 
                'message': 'Invalid date format. Use DD/MM/YYYY'
            }, status=400)
        
        # Check if PO already exists
        if PurchaseOrder.objects.filter(po_id=po_id).exists():
            return JsonResponse({
                'success': False, 
                'message': 'Purchase Order ID already exists'
            }, status=400)
        
        # Get foreign key objects
        try:
            supplier = Supplier.objects.get(supplier_id=supplier_id)
            county = County.objects.get(county=county_name) if county_name else None
            town = Town.objects.get(town=town_name) if town_name else None
        except (Supplier.DoesNotExist, County.DoesNotExist, Town.DoesNotExist) as e:
            return JsonResponse({
                'success': False, 
                'message': f'Invalid reference: {str(e)}'
            }, status=400)
        
        # Start transaction
        with transaction.atomic():
            # Calculate total amount
            total_amount = sum(Decimal(str(item['total_purchase_price'])) for item in items)
            
            # Create Purchase Order with initial amount_paid = 0
            purchase_order = PurchaseOrder.objects.create(
                po_id=po_id,
                date=po_date,
                supplier_id=supplier,
                supplier_name=supplier_name,
                bill_number=bill_number,
                county=county,
                town=town,
                total_amount=total_amount,
                amount_paid=Decimal('0.00'),
                payment_status=None,  # Will be set by update_purchase_order_statuses
                shipping_status=None  # Will be set by update_purchase_order_statuses
            )
            
            # Create Purchase Details and update inventory
            for item_data in items:
                detail_date = parse_date(item_data['date'])
                if not detail_date:
                    detail_date = po_date
                
                # Get inventory item
                try:
                    inventory_item = Inventory.objects.get(item_id=item_data['item_id'])
                except Inventory.DoesNotExist:
                    raise Exception(f"Inventory item {item_data['item_id']} not found")
                
                # Create purchase detail
                PurchaseDetail.objects.create(
                    detail_id=item_data['detail_id'],
                    po_id=purchase_order,
                    date=detail_date,
                    supplier_id=supplier,
                    supplier_name=supplier_name,
                    county=county,
                    town=town,
                    bill_number=bill_number,
                    item_id=inventory_item,
                    item_type=item_data['item_type'],
                    item_category=item_data['item_category'],
                    item_subcategory=item_data['item_subcategory'],
                    item_name=item_data['item_name'],
                    quantity_purchased=item_data['quantity_purchased'],
                    unit_cost=Decimal(str(item_data['unit_cost'])),
                    tax_rate=Decimal(str(item_data['tax_rate']))
                )
                
                # Update inventory quantities
                inventory_item.quantity_purchased += item_data['quantity_purchased']
                inventory_item.save()
            
            # Update supplier total purchases
            supplier.total_purchases += total_amount
            supplier.save()
            
            # *** AUTOMATICALLY SET PAYMENT AND SHIPPING STATUS ***
            status_info = update_purchase_order_statuses(purchase_order)
        
        return JsonResponse({
            'success': True, 
            'message': 'Purchase Order created successfully',
            'status_info': status_info
        })
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

# ========================== UPDATING THE PURCHASE ORDER WHENEVER ANY MODIFICATION IS MADE ===========================================================================
@csrf_exempt
@login_required(login_url='/login/')
def api_update_purchase_order(request):
    """
    Update existing purchase order with proper Detail ID handling
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
    
    try:
        data = json.loads(request.body)
        
        po_id = data.get('po_id', '').strip()
        items = data.get('items', [])
        
        if not po_id:
            return JsonResponse({'success': False, 'message': 'PO ID is required'}, status=400)
        
        # Get existing PO
        try:
            purchase_order = PurchaseOrder.objects.get(po_id=po_id)
        except PurchaseOrder.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Purchase order not found'}, status=404)
        
        with transaction.atomic():
            # Get existing details
            existing_details = {d.detail_id: d for d in PurchaseDetail.objects.filter(po_id=purchase_order)}
            
            new_total = Decimal('0.00')
            processed_details = set()
            
            for item_data in items:
                detail_id = item_data['detail_id']
                processed_details.add(detail_id)
                
                new_qty = item_data['quantity_purchased']
                total_price = Decimal(str(item_data['total_purchase_price']))
                new_total += total_price
                
                # Get inventory item
                inventory_item = Inventory.objects.get(item_id=item_data['item_id'])
                
                if detail_id in existing_details:
                    # ✅ UPDATE EXISTING DETAIL
                    detail = existing_details[detail_id]
                    old_qty = detail.quantity_purchased
                    qty_diff = new_qty - old_qty
                    
                    # Update detail fields
                    detail.quantity_purchased = new_qty
                    detail.unit_cost = Decimal(str(item_data['unit_cost']))
                    detail.tax_rate = Decimal(str(item_data['tax_rate']))
                    detail.item_type = item_data['item_type']
                    detail.item_category = item_data['item_category']
                    detail.item_subcategory = item_data['item_subcategory']
                    detail.item_name = item_data['item_name']
                    
                    # Handle item change
                    if detail.item_id.item_id != item_data['item_id']:
                        # Reverse old item
                        detail.item_id.quantity_purchased -= old_qty
                        detail.item_id.save()
                        
                        # Add to new item
                        inventory_item.quantity_purchased += new_qty
                        inventory_item.save()
                        
                        detail.item_id = inventory_item
                    else:
                        # Same item, adjust quantity
                        inventory_item.quantity_purchased += qty_diff
                        inventory_item.save()
                    
                    detail.save()
                    print(f"✅ UPDATED Detail: {detail_id}")
                    
                else:
                    # ✅ CREATE NEW DETAIL (this was missing proper creation)
                    detail_date = parse_date(item_data.get('date', ''))
                    if not detail_date:
                        detail_date = purchase_order.date
                    
                    PurchaseDetail.objects.create(
                        detail_id=detail_id,
                        po_id=purchase_order,
                        date=detail_date,
                        supplier_id=purchase_order.supplier_id,
                        supplier_name=purchase_order.supplier_name,
                        county=purchase_order.county,
                        town=purchase_order.town,
                        bill_number=purchase_order.bill_number,
                        item_id=inventory_item,
                        item_type=item_data['item_type'],
                        item_category=item_data['item_category'],
                        item_subcategory=item_data['item_subcategory'],
                        item_name=item_data['item_name'],
                        quantity_purchased=new_qty,
                        unit_cost=Decimal(str(item_data['unit_cost'])),
                        tax_rate=Decimal(str(item_data['tax_rate']))
                    )
                    
                    # Update inventory
                    inventory_item.quantity_purchased += new_qty
                    inventory_item.save()
                    
                    print(f"✅ CREATED NEW Detail: {detail_id}")
            
            # ✅ DELETE REMOVED DETAILS
            for detail_id, detail in existing_details.items():
                if detail_id not in processed_details:
                    # Reverse inventory
                    detail.item_id.quantity_purchased -= detail.quantity_purchased
                    detail.item_id.save()
                    detail.delete()
                    print(f"✅ DELETED Detail: {detail_id}")
            
            # Update PO total
            old_total = purchase_order.total_amount
            total_diff = new_total - old_total
            
            purchase_order.total_amount = new_total
            purchase_order.save()
            
            # Update supplier total
            supplier = purchase_order.supplier_id
            supplier.total_purchases += total_diff
            supplier.save()
            
            # Update statuses
            status_info = update_purchase_order_statuses(purchase_order)
        
        return JsonResponse({
            'success': True, 
            'message': 'Purchase Order updated successfully',
            'status_info': status_info
        })
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        print(f"❌ UPDATE ERROR: {str(e)}")
        return JsonResponse({'success': False, 'message': str(e)}, status=500)
    

@csrf_exempt
@login_required(login_url='/login/')
def api_delete_purchase_detail(request):
    """Delete a purchase detail and update related records"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
    
    try:
        data = json.loads(request.body)
        detail_id = data.get('detail_id', '').strip()
        
        if not detail_id:
            return JsonResponse({'success': False, 'message': 'Detail ID is required'}, status=400)
        
        try:
            detail = PurchaseDetail.objects.get(detail_id=detail_id)
        except PurchaseDetail.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Purchase detail not found'}, status=404)
        
        with transaction.atomic():
            # Reverse inventory quantity
            inventory_item = detail.item_id
            inventory_item.quantity_purchased -= detail.quantity_purchased
            inventory_item.save()
            
            # Update PO total
            po = detail.po_id
            po.total_amount -= detail.total_purchase_price
            po.save()
            
            # Update supplier total
            supplier = detail.supplier_id
            supplier.total_purchases -= detail.total_purchase_price
            supplier.save()
            
            # Delete detail
            detail.delete()
        
        return JsonResponse({
            'success': True, 
            'message': 'Purchase detail deleted successfully'
        })
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
    
# ==================== UPDATING THE SHIPPING AND PAYMENT STATUSES FOR THE APPLICATION ========================================================================================

def calculate_payment_status(total_amount, amount_paid):
    """
    Calculate payment status based on payment percentage
    
    Args:
        total_amount (Decimal): Total amount of the purchase order
        amount_paid (Decimal): Amount already paid
    
    Returns:
        str: Payment status ('PENDING', 'PARTIAL PAYMENT', 'COMPLETED')
    """
    total_amount = Decimal(str(total_amount))
    amount_paid = Decimal(str(amount_paid))
    
    # If no payment made
    if amount_paid == Decimal('0.00'):
        return 'PENDING'
    
    # If fully paid
    elif amount_paid >= total_amount:
        return 'COMPLETED'
    
    # If partially paid
    else:
        return 'PARTIAL PAYMENT'


def calculate_shipping_status(total_amount, amount_paid):
    """
    Calculate shipping status based on payment percentage
    
    Payment Percentage -> Shipping Status:
    - 0% -> PENDING
    - 1%-49% -> PROCESSING
    - 50%-74% -> DISPATCHED
    - 75%-99% -> IN TRANSIT
    - 100% -> DELIVERED
    
    Args:
        total_amount (Decimal): Total amount of the purchase order
        amount_paid (Decimal): Amount already paid
    
    Returns:
        str: Shipping status
    """
    total_amount = Decimal(str(total_amount))
    amount_paid = Decimal(str(amount_paid))
    
    # If no payment made
    if amount_paid == Decimal('0.00'):
        return 'PENDING'
    
    # Calculate payment percentage
    payment_percentage = (amount_paid / total_amount) * 100
    
    # Determine shipping status based on percentage
    if payment_percentage >= 100:
        return 'DELIVERED'
    elif payment_percentage >= 75:
        return 'IN TRANSIT'
    elif payment_percentage >= 50:
        return 'DISPATCHED'
    elif payment_percentage >= 1:
        return 'PROCESSING'
    else:
        return 'PENDING'


def update_purchase_order_statuses(purchase_order):
    """
    Update both payment and shipping status for a purchase order
    
    Args:
        purchase_order: PurchaseOrder instance
    
    Returns:
        dict: Updated status information
    """
    # Get or create status objects
    payment_status_name = calculate_payment_status(
        purchase_order.total_amount, 
        purchase_order.amount_paid
    )
    
    shipping_status_name = calculate_shipping_status(
        purchase_order.total_amount, 
        purchase_order.amount_paid
    )
    
    # Get or create PaymentStatus and ShippingStatus objects
    payment_status, _ = PaymentStatus.objects.get_or_create(
        payment_status=payment_status_name
    )
    
    shipping_status, _ = ShippingStatus.objects.get_or_create(
        shipping_status=shipping_status_name
    )
    
    # Update the purchase order
    purchase_order.payment_status = payment_status
    purchase_order.shipping_status = shipping_status
    purchase_order.save()
    
    return {
        'payment_status': payment_status_name,
        'shipping_status': shipping_status_name,
        'payment_percentage': float((purchase_order.amount_paid / purchase_order.total_amount) * 100),
        'balance_remaining': float(purchase_order.total_amount - purchase_order.amount_paid)
    }   
# ============================================ SALES MODULE VIEWS ======================================================================================

@login_required(login_url='/login/')
def sales_content(request):
    """Load Sales content page"""
    return render(request, 'Sales.html')

# ================================================ RECEIPT STATUS APIs =====================================================================================

@csrf_exempt
@login_required(login_url='/login/')
def api_get_receipt_statuses(request):
    """Get all receipt statuses"""
    try:
        statuses = list(ReceiptStatus.objects.values_list('receipt_status', flat=True).order_by('receipt_status'))
        return JsonResponse({'success': True, 'data': statuses})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

@csrf_exempt
@login_required(login_url='/login/')
def api_add_receipt_status(request):
    """Add new receipt status"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
    
    try:
        data = json.loads(request.body)
        status_name = data.get('status_name', '').strip()
        
        if not status_name:
            return JsonResponse({'success': False, 'message': 'Receipt status name is required'}, status=400)
        
        if ReceiptStatus.objects.filter(receipt_status__iexact=status_name).exists():
            return JsonResponse({'success': False, 'message': 'Receipt status already exists'}, status=400)
        
        ReceiptStatus.objects.create(receipt_status=status_name)
        
        return JsonResponse({'success': True, 'message': 'Receipt status added successfully'})
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

# ======================== SHIPPING STATUS APIs (BASICALLY REUSED FROM THE PURCHASES MODULE) =======================================================================================

@csrf_exempt
@login_required(login_url='/login/')
def api_get_shipping_statuses_sales(request):
    """Get all shipping statuses"""
    try:
        statuses = list(ShippingStatus.objects.values_list('shipping_status', flat=True).order_by('shipping_status'))
        return JsonResponse({'success': True, 'data': statuses})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

# =================================== SALES ORDER APIs ===================================================================================================

@csrf_exempt
@login_required(login_url='/login/')
def api_generate_so_id(request):
    """Generate unique SO ID in format SO00001 and corresponding Invoice Number I00001"""
    try:
        max_so = SalesOrder.objects.aggregate(Max('so_id'))['so_id__max']
        
        if max_so:
            num_part = int(max_so[2:]) + 1
        else:
            num_part = 1
        
        new_so_id = f"SO{num_part:05d}"
        
        # Make sure it doesn't exist
        while SalesOrder.objects.filter(so_id=new_so_id).exists():
            num_part += 1
            new_so_id = f"SO{num_part:05d}"
        
        # Generate corresponding invoice number
        invoice_number = f"I{num_part:05d}"
        
        return JsonResponse({
            'success': True, 
            'so_id': new_so_id,
            'invoice_number': invoice_number
        })
    
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

# ======================= GENERATING THE SALES DETAIL ID =====================================================================================================================
@csrf_exempt
@login_required(login_url='/login/')
def api_generate_sales_detail_id(request):
    """
    Generate unique Detail ID for SALES in format SD00001
    Each call generates a FRESH unique ID
    """
    try:
        # Always fetch fresh from database
        max_sales_detail = SalesDetail.objects.aggregate(Max('detail_id'))['detail_id__max']
        
        if max_sales_detail:
            # Extract numeric part (skip 'SD' prefix)
            num_part = int(max_sales_detail[2:]) + 1
        else:
            # First sales detail ever
            num_part = 1
        
        # Format as SD00001
        new_detail_id = f"SD{num_part:05d}"
        
        # Safety check: ensure it doesn't exist
        attempts = 0
        while SalesDetail.objects.filter(detail_id=new_detail_id).exists():
            num_part += 1
            new_detail_id = f"SD{num_part:05d}"
            attempts += 1
            if attempts > 1000:
                return JsonResponse({'success': False, 'message': 'Could not generate unique Detail ID'}, status=500)
        
        print(f"✅ GENERATED SALES DETAIL ID: {new_detail_id} at {time.strftime('%H:%M:%S')}")
        
        return JsonResponse({'success': True, 'detail_id': new_detail_id})
    
    except Exception as e:
        print(f"❌ ERROR generating Detail ID: {str(e)}")
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

@csrf_exempt
@login_required(login_url='/login/')
def api_get_next_sales_detail_number(request):  
    """
    Get the next available detail number for Sales
    Returns just the number, not the full ID
    """
    try:
        max_sales_detail = SalesDetail.objects.aggregate(Max('detail_id'))['detail_id__max']
        
        if max_sales_detail:
            # Extract number from SD00007 -> 7, then add 1
            next_number = int(max_sales_detail[2:]) + 1
        else:
            next_number = 1
        
        print(f"📊 Next Sales Detail Number: {next_number}")
        return JsonResponse({'success': True, 'next_number': next_number})
    
    except Exception as e:
        print(f"❌ Error getting next detail number: {str(e)}")
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
@login_required(login_url='/login/')
def api_get_sales_orders(request):
    """Get all sales orders with calculated totals"""
    try:
        sales_orders = SalesOrder.objects.select_related(
            'customer_id', 'county', 'town', 'receipt_status', 'shipping_status'
        ).all().order_by('-date')
        
        so_list = []
        for so in sales_orders:
            # Format date as DD/MM/YYYY
            date_str = so.date.strftime('%d/%m/%Y') if so.date else ''
            
            so_list.append({
                'so_id': so.so_id,
                'date': date_str,
                'customer_id': so.customer_id.customer_id if so.customer_id else '',
                'customer_name': so.customer_name,
                'invoice_number': so.invoice_number,
                'county': so.county.county if so.county else '',
                'town': so.town.town if so.town else '',
                'total_amount': float(so.total_amount),
                'amount_received': float(so.amount_received),
                'balance_left': float(so.total_amount - so.amount_received),
                'receipt_status': so.receipt_status.receipt_status if so.receipt_status else '',
                'shipping_status': so.shipping_status.shipping_status if so.shipping_status else ''
            })
        
        return JsonResponse({'success': True, 'data': so_list})
    
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

@csrf_exempt
@login_required(login_url='/login/')
def api_get_so_details(request, so_id):
    """Get sales order details including all line items"""
    try:
        # Get sales order
        so = SalesOrder.objects.select_related(
            'customer_id', 'county', 'town', 'receipt_status', 'shipping_status'
        ).get(so_id=so_id)
        
        # Format date
        date_str = so.date.strftime('%d/%m/%Y') if so.date else ''
        
        so_data = {
            'so_id': so.so_id,
            'date': date_str,
            'customer_id': so.customer_id.customer_id if so.customer_id else '',
            'customer_name': so.customer_name,
            'invoice_number': so.invoice_number,
            'county': so.county.county if so.county else '',
            'town': so.town.town if so.town else '',
            'total_amount': float(so.total_amount),
            'amount_received': float(so.amount_received),
            'receipt_status': so.receipt_status.receipt_status if so.receipt_status else '',
            'shipping_status': so.shipping_status.shipping_status if so.shipping_status else ''
        }
        
        # Get sales details
        details = SalesDetail.objects.filter(so_id=so).order_by('detail_id')
        details_list = []
        
        for detail in details:
            detail_date_str = detail.date.strftime('%d/%m/%Y') if detail.date else ''
            
            details_list.append({
                'detail_id': detail.detail_id,
                'date': detail_date_str,
                'so_id': detail.so_id.so_id,
                'customer_id': detail.customer_id.customer_id if detail.customer_id else '',
                'customer_name': detail.customer_name,
                'county': detail.county.county if detail.county else '',
                'town': detail.town.town if detail.town else '',
                'invoice_number': detail.invoice_number,
                'item_id': detail.item_id.item_id if detail.item_id else '',
                'item_type': detail.item_type,
                'item_category': detail.item_category,
                'item_subcategory': detail.item_subcategory,
                'item_name': detail.item_name,
                'quantity_sold': detail.quantity_sold,
                'unit_price': float(detail.unit_price),
                'tax_rate': float(detail.tax_rate),
                'price_excluding_tax': float(detail.price_excluding_tax),
                'total_tax': float(detail.total_tax),
                'price_including_tax': float(detail.price_including_tax),
                'shipping_fees': float(detail.shipping_fees),
                'total_sales_price': float(detail.total_sales_price)
            })
        
        return JsonResponse({
            'success': True,
            'data': {
                'sales_order': so_data,
                'sales_details': details_list
            }
        })
    
    except SalesOrder.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Sales order not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

# =========================== ADDING A NEW SALES ORDER ======================================================================================================================================
@csrf_exempt
@login_required(login_url='/login/')
def api_add_sales_order(request):
    """Add new sales order with details - DECREASES inventory quantities"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
    
    try:
        data = json.loads(request.body)
        
        # Extract SO header data
        so_id = data.get('so_id', '').strip()
        date_str = data.get('date', '').strip()
        customer_id = data.get('customer_id', '').strip()
        customer_name = data.get('customer_name', '').strip()
        county_name = data.get('county', '').strip()
        town_name = data.get('town', '').strip()
        invoice_number = data.get('invoice_number', '').strip()
        items = data.get('items', [])
        
        # Validation
        if not all([so_id, date_str, customer_name, invoice_number]):
            return JsonResponse({
                'success': False, 
                'message': 'All SO header fields are required'
            }, status=400)
        
        if not items:
            return JsonResponse({
                'success': False, 
                'message': 'At least one item is required'
            }, status=400)
        
        # Parse date
        so_date = parse_date(date_str)
        if not so_date:
            return JsonResponse({
                'success': False, 
                'message': 'Invalid date format. Use DD/MM/YYYY'
            }, status=400)
        
        # Check if SO already exists
        if SalesOrder.objects.filter(so_id=so_id).exists():
            return JsonResponse({
                'success': False, 
                'message': 'Sales Order ID already exists'
            }, status=400)
        
        # Get foreign key objects
        try:
            customer = Customer.objects.get(customer_id=customer_id)
            county = County.objects.get(county=county_name) if county_name else None
            town = Town.objects.get(town=town_name) if town_name else None
        except (Customer.DoesNotExist, County.DoesNotExist, Town.DoesNotExist) as e:
            return JsonResponse({
                'success': False, 
                'message': f'Invalid reference: {str(e)}'
            }, status=400)
        
        # Start transaction
        with transaction.atomic():
            # Calculate total amount
            total_amount = sum(Decimal(str(item['total_sales_price'])) for item in items)
            
            # Create Sales Order
            sales_order = SalesOrder.objects.create(
                so_id=so_id,
                date=so_date,
                customer_id=customer,
                customer_name=customer_name,
                invoice_number=invoice_number,
                county=county,
                town=town,
                total_amount=total_amount,
                amount_received=Decimal('0.00'),
                receipt_status=None,  # Will be set by update_sales_order_statuses
                shipping_status=None  # Will be set by update_sales_order_statuses
            )
            
            # Create Sales Details and update inventory
            for item_data in items:
                detail_date = parse_date(item_data['date'])
                if not detail_date:
                    detail_date = so_date
                
                # Get inventory item
                try:
                    inventory_item = Inventory.objects.get(item_id=item_data['item_id'])
                except Inventory.DoesNotExist:
                    raise Exception(f"Inventory item {item_data['item_id']} not found")
                
                # Check if sufficient stock exists
                qty_to_sell = item_data['quantity_sold']
                available_qty = inventory_item.quantity_purchased - inventory_item.quantity_sold
                
                if qty_to_sell > available_qty:
                    raise Exception(f"Insufficient stock for {item_data['item_name']}. Available: {available_qty}, Requested: {qty_to_sell}")
                
                # Create sales detail
                SalesDetail.objects.create(
                    detail_id=item_data['detail_id'],
                    so_id=sales_order,
                    date=detail_date,
                    customer_id=customer,
                    customer_name=customer_name,
                    county=county,
                    town=town,
                    invoice_number=invoice_number,
                    item_id=inventory_item,
                    item_type=item_data['item_type'],
                    item_category=item_data['item_category'],
                    item_subcategory=item_data['item_subcategory'],
                    item_name=item_data['item_name'],
                    quantity_sold=qty_to_sell,
                    unit_price=Decimal(str(item_data['unit_price'])),
                    tax_rate=Decimal(str(item_data['tax_rate']))
                )
                
                # Update inventory quantities (DECREASE quantity_sold)
                inventory_item.quantity_sold += qty_to_sell
                inventory_item.save()  # This will trigger reorder check
            
            # Update customer total sales
            customer.total_sales += total_amount
            customer.save()
            
            # AUTOMATICALLY SET RECEIPT AND SHIPPING STATUS 
            status_info = update_sales_order_statuses(sales_order)
        
        return JsonResponse({
            'success': True, 
            'message': 'Sales Order created successfully',
            'status_info': status_info
        })
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

# ============================ CALCULATING THE RECEIPT STATUS IN THE SALES ORDER ===============================================================================
@csrf_exempt
@login_required(login_url='/login/')
def calculate_receipt_status(total_amount, amount_received):
    """
    Calculate receipt status based on payment percentage
    
    Args:
        total_amount (Decimal): Total amount of the sales order
        amount_received (Decimal): Amount already received
    
    Returns:
        str: Receipt status ('PENDING', 'PARTIAL PAYMENT', 'COMPLETED')
    """
    total_amount = Decimal(str(total_amount))
    amount_received = Decimal(str(amount_received))
    
    # If no payment received
    if amount_received == Decimal('0.00'):
        return 'PENDING'
    
    # If fully paid
    elif amount_received >= total_amount:
        return 'COMPLETED'
    
    # If partially paid
    else:
        return 'PARTIAL PAYMENT'
    
# ============================ CALCULATING THE SHIPPING STATUS IN THE SALES ORDER ================================================================================
def calculate_sales_shipping_status(total_amount, amount_received):
    """
    Calculate shipping status based on payment percentage
    
    Payment Percentage -> Shipping Status:
    - 0% -> PENDING
    - 1%-49% -> PROCESSING
    - 50%-74% -> DISPATCHED
    - 75%-99% -> IN TRANSIT
    - 100% -> DELIVERED
    
    Args:
        total_amount (Decimal): Total amount of the sales order
        amount_received (Decimal): Amount already received
    
    Returns:
        str: Shipping status
    """
    total_amount = Decimal(str(total_amount))
    amount_received = Decimal(str(amount_received))
    
    # If no payment received
    if amount_received == Decimal('0.00'):
        return 'PENDING'
    
    # Calculate payment percentage
    payment_percentage = (amount_received / total_amount) * 100
    
    # Determine shipping status based on percentage
    if payment_percentage >= 100:
        return 'DELIVERED'
    elif payment_percentage >= 75:
        return 'IN TRANSIT'
    elif payment_percentage >= 50:
        return 'DISPATCHED'
    elif payment_percentage >= 1:
        return 'PROCESSING'
    else:
        return 'PENDING'


def update_sales_order_statuses(sales_order):
    """
    Update both receipt and shipping status for a sales order
    
    Args:
        sales_order: SalesOrder instance
    
    Returns:
        dict: Updated status information
    """
    # Calculate status names
    receipt_status_name = calculate_receipt_status(
        sales_order.total_amount, 
        sales_order.amount_received
    )
    
    shipping_status_name = calculate_sales_shipping_status(
        sales_order.total_amount, 
        sales_order.amount_received
    )
    
    # Get or create ReceiptStatus and ShippingStatus objects
    receipt_status, _ = ReceiptStatus.objects.get_or_create(
        receipt_status=receipt_status_name
    )
    
    shipping_status, _ = ShippingStatus.objects.get_or_create(
        shipping_status=shipping_status_name
    )
    
    # Update the sales order
    sales_order.receipt_status = receipt_status
    sales_order.shipping_status = shipping_status
    sales_order.save()
    
    return {
        'receipt_status': receipt_status_name,
        'shipping_status': shipping_status_name,
        'payment_percentage': float((sales_order.amount_received / sales_order.total_amount) * 100) if sales_order.total_amount > 0 else 0,
        'balance_remaining': float(sales_order.total_amount - sales_order.amount_received)
    }
    
# ================================== UPDATING THE SALES ORDER ==============================================================================================================================    
@csrf_exempt
@login_required(login_url='/login/')
def api_update_sales_order(request):
    """Update existing sales order"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
    
    try:
        data = json.loads(request.body)
        
        so_id = data.get('so_id', '').strip()
        items = data.get('items', [])
        
        if not so_id:
            return JsonResponse({'success': False, 'message': 'SO ID is required'}, status=400)
        
        # Get existing SO
        try:
            sales_order = SalesOrder.objects.get(so_id=so_id)
        except SalesOrder.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Sales order not found'}, status=404)
        
        with transaction.atomic():
            # Get existing details to calculate old quantities
            existing_details = {d.detail_id: d for d in SalesDetail.objects.filter(so_id=sales_order)}
            
            # Process each item
            new_total = Decimal('0.00')
            processed_details = set()
            
            for item_data in items:
                detail_id = item_data['detail_id']
                processed_details.add(detail_id)
                
                new_qty = item_data['quantity_sold']
                total_price = Decimal(str(item_data['total_sales_price']))
                new_total += total_price
                
                if detail_id in existing_details:
                    # Update existing detail
                    detail = existing_details[detail_id]
                    old_qty = detail.quantity_sold
                    qty_diff = new_qty - old_qty
                    
                    # Update detail
                    detail.quantity_sold = new_qty
                    detail.unit_price = Decimal(str(item_data['unit_price']))
                    detail.tax_rate = Decimal(str(item_data['tax_rate']))
                    detail.item_type = item_data['item_type']
                    detail.item_category = item_data['item_category']
                    detail.item_subcategory = item_data['item_subcategory']
                    detail.item_name = item_data['item_name']
                    
                    # Update item_id if changed
                    new_item = Inventory.objects.get(item_id=item_data['item_id'])
                    if detail.item_id != new_item:
                        # Reverse old item quantity (add back)
                        detail.item_id.quantity_sold -= old_qty
                        detail.item_id.save()
                        
                        # Deduct from new item
                        new_item.quantity_sold += new_qty
                        new_item.save()
                        
                        detail.item_id = new_item
                    else:
                        # Same item, adjust quantity
                        detail.item_id.quantity_sold += qty_diff
                        detail.item_id.save()
                    
                    detail.save()
            
            # Handle deleted items (items in DB but not in update)
            for detail_id, detail in existing_details.items():
                if detail_id not in processed_details:
                    # This item was deleted - reverse inventory (add back)
                    detail.item_id.quantity_sold -= detail.quantity_sold
                    detail.item_id.save()
                    detail.delete()
            
            # Update SO total
            old_total = sales_order.total_amount
            total_diff = new_total - old_total
            
            sales_order.total_amount = new_total
            sales_order.save()
            
            # Update customer total sales
            customer = sales_order.customer_id
            customer.total_sales += total_diff
            customer.save()
            
            # *** UPDATE STATUSES ***
            status_info = update_sales_order_statuses(sales_order)
        
        return JsonResponse({
            'success': True, 
            'message': 'Sales Order updated successfully',
            'status_info': status_info
        })
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

# ================================== DELETING THE SALES ORDER ======================================================================================================================
@csrf_exempt
@login_required(login_url='/login/')
def api_delete_sales_detail(request):
    """Delete a sales detail and update related records"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
    
    try:
        data = json.loads(request.body)
        detail_id = data.get('detail_id', '').strip()
        
        if not detail_id:
            return JsonResponse({'success': False, 'message': 'Detail ID is required'}, status=400)
        
        try:
            detail = SalesDetail.objects.get(detail_id=detail_id)
        except SalesDetail.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Sales detail not found'}, status=404)
        
        with transaction.atomic():
            # Reverse inventory quantity (add back)
            inventory_item = detail.item_id
            inventory_item.quantity_sold -= detail.quantity_sold
            inventory_item.save()
            
            # Update SO total
            so = detail.so_id
            so.total_amount -= detail.total_sales_price
            so.save()
            
            # Update customer total
            customer = detail.customer_id
            customer.total_sales -= detail.total_sales_price
            customer.save()
            
            # Delete detail
            detail.delete()
        
        return JsonResponse({
            'success': True, 
            'message': 'Sales detail deleted successfully'
        })
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
    
# ============================== NEW API ENDPOINT FOR THE PAYMENTS MODULE ===============================================================================
@csrf_exempt
@login_required(login_url='/login/')
def api_record_payment(request):
    """
    Record a payment and automatically update PO statuses
    This endpoint will be called from the Payments module
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
    
    try:
        data = json.loads(request.body)
        
        po_id = data.get('po_id', '').strip()
        payment_amount = Decimal(str(data.get('payment_amount', 0)))
        payment_mode = data.get('payment_mode', '').strip()
        transaction_id = data.get('transaction_id', '').strip()
        payment_date_str = data.get('payment_date', '').strip()
        
        # Validation
        if not all([po_id, payment_amount, payment_mode, transaction_id]):
            return JsonResponse({
                'success': False, 
                'message': 'All payment fields are required'
            }, status=400)
        
        if payment_amount <= 0:
            return JsonResponse({
                'success': False, 
                'message': 'Payment amount must be greater than zero'
            }, status=400)
        
        # Get purchase order
        try:
            purchase_order = PurchaseOrder.objects.get(po_id=po_id)
        except PurchaseOrder.DoesNotExist:
            return JsonResponse({
                'success': False, 
                'message': 'Purchase order not found'
            }, status=404)
        
        # Check if payment exceeds balance
        balance = purchase_order.total_amount - purchase_order.amount_paid
        if payment_amount > balance:
            return JsonResponse({
                'success': False, 
                'message': f'Payment amount ({payment_amount}) exceeds remaining balance ({balance})'
            }, status=400)
        
        # Parse payment date
        payment_date = parse_date(payment_date_str)
        if not payment_date:
            return JsonResponse({
                'success': False, 
                'message': 'Invalid date format. Use DD/MM/YYYY'
            }, status=400)
        
        with transaction.atomic():
            # Get payment mode object
            pmt_mode, _ = PaymentMode.objects.get_or_create(payment_mode=payment_mode)
            
            # Create payment record
            Payment.objects.create(
                transaction_id=transaction_id,
                date=payment_date,
                supplier_id=purchase_order.supplier_id,
                supplier_name=purchase_order.supplier_name,
                county=purchase_order.county,
                town=purchase_order.town,
                po_id=purchase_order,
                bill_number=purchase_order.bill_number,
                payment_mode=pmt_mode,
                amount_paid=payment_amount
            )
            
            # Update purchase order amount_paid
            purchase_order.amount_paid += payment_amount
            purchase_order.save()
            
            # Update supplier total_payments
            supplier = purchase_order.supplier_id
            supplier.total_payments += payment_amount
            supplier.save()
            
            # AUTOMATICALLY UPDATE PAYMENT AND SHIPPING STATUS
            status_info = update_purchase_order_statuses(purchase_order)
        
        return JsonResponse({
            'success': True, 
            'message': 'Payment recorded successfully',
            'status_info': status_info,
            'new_balance': float(purchase_order.total_amount - purchase_order.amount_paid)
        })
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


# ============================================
# ENDPOINT TO MANUALLY RECALCULATE ALL STATUSES
# ============================================

@csrf_exempt
@login_required(login_url='/login/')
def api_recalculate_all_po_statuses(request):
    """
    Recalculate statuses for all purchase orders
    Useful for one-time migration or fixing data inconsistencies
    """
    try:
        purchase_orders = PurchaseOrder.objects.all()
        updated_count = 0
        
        for po in purchase_orders:
            update_purchase_order_statuses(po)
            updated_count += 1
        
        return JsonResponse({
            'success': True,
            'message': f'Successfully updated statuses for {updated_count} purchase orders'
        })
    
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


# ============================================
# PAYMENTS MODULE VIEWS
# ============================================

@login_required(login_url='/login/')
def payments_content(request):
    """Load Payments content page"""
    return render(request, 'Payments.html')


# ============================================
# PAYMENT MODE APIs
# ============================================

@csrf_exempt
@login_required(login_url='/login/')
def api_get_payment_modes(request):
    """Get all payment modes"""
    try:
        modes = list(PaymentMode.objects.values_list('payment_mode', flat=True).order_by('payment_mode'))
        return JsonResponse({'success': True, 'data': modes})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
@login_required(login_url='/login/')
def api_add_payment_mode(request):
    """Add new payment mode"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
    
    try:
        data = json.loads(request.body)
        mode_name = data.get('mode_name', '').strip()
        
        if not mode_name:
            return JsonResponse({'success': False, 'message': 'Payment mode name is required'}, status=400)
        
        if PaymentMode.objects.filter(payment_mode__iexact=mode_name).exists():
            return JsonResponse({'success': False, 'message': 'Payment mode already exists'}, status=400)
        
        PaymentMode.objects.create(payment_mode=mode_name)
        
        return JsonResponse({'success': True, 'message': 'Payment mode added successfully'})
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


# ============================================
# GENERATE TRANSACTION ID
# ============================================

@csrf_exempt
@login_required(login_url='/login/')
def api_generate_transaction_id(request):
    """Generate unique Transaction ID in format TRANSPAY00001"""
    try:
        max_payment = Payment.objects.aggregate(Max('transaction_id'))['transaction_id__max']
        
        if max_payment:
            # Extract numeric part from TRANSPAY00001
            num_part = int(max_payment[8:]) + 1
        else:
            num_part = 1
        
        new_id = f"TRANSPAY{num_part:05d}"
        
        # Safety check
        while Payment.objects.filter(transaction_id=new_id).exists():
            num_part += 1
            new_id = f"TRANSPAY{num_part:05d}"
        
        return JsonResponse({'success': True, 'transaction_id': new_id})
    
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


# ============================================
# GET ALL PAYMENTS
# ============================================

@csrf_exempt
@login_required(login_url='/login/')
def api_get_payments(request):
    """Get all payments with full details"""
    try:
        payments = Payment.objects.select_related(
            'supplier_id', 'county', 'town', 'po_id', 'payment_mode'
        ).all().order_by('-date')
        
        payments_list = []
        for payment in payments:
            # Format date as DD/MM/YYYY
            date_str = payment.date.strftime('%d/%m/%Y') if payment.date else ''
            
            payments_list.append({
                'transaction_id': payment.transaction_id,
                'date': date_str,
                'supplier_id': payment.supplier_id.supplier_id if payment.supplier_id else '',
                'supplier_name': payment.supplier_name,
                'county': payment.county.county if payment.county else '',
                'town': payment.town.town if payment.town else '',
                'po_id': payment.po_id.po_id if payment.po_id else '',
                'bill_number': payment.bill_number,
                'payment_mode': payment.payment_mode.payment_mode if payment.payment_mode else '',
                'amount_paid': float(payment.amount_paid)
            })
        
        return JsonResponse({'success': True, 'data': payments_list})
    
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


# ============================================
# HELPER: Calculate and Update PO Statuses
# ============================================

def update_purchase_order_statuses(purchase_order):
    """
    Update both payment and shipping status for a purchase order
    
    Payment Status Rules:
    - 0% paid -> PENDING
    - 1-99% paid -> PARTIAL PAYMENT
    - 100% paid -> COMPLETED
    
    Shipping Status Rules:
    - 0% paid -> PENDING
    - 1-49% paid -> PROCESSING
    - 50-74% paid -> DISPATCHED
    - 75-99% paid -> IN TRANSIT
    - 100% paid -> DELIVERED
    """
    total_amount = Decimal(str(purchase_order.total_amount))
    amount_paid = Decimal(str(purchase_order.amount_paid))
    
    # Calculate payment percentage
    if total_amount > 0:
        payment_percentage = (amount_paid / total_amount) * 100
    else:
        payment_percentage = 0
    
    # Determine Payment Status
    if amount_paid == Decimal('0.00'):
        payment_status_name = 'PENDING'
    elif amount_paid >= total_amount:
        payment_status_name = 'COMPLETED'
    else:
        payment_status_name = 'PARTIAL PAYMENT'
    
    # Determine Shipping Status
    if payment_percentage == 0:
        shipping_status_name = 'PENDING'
    elif payment_percentage >= 100:
        shipping_status_name = 'DELIVERED'
    elif payment_percentage >= 75:
        shipping_status_name = 'IN TRANSIT'
    elif payment_percentage >= 50:
        shipping_status_name = 'DISPATCHED'
    else:
        shipping_status_name = 'PROCESSING'
    
    # Get or create status objects
    payment_status, _ = PaymentStatus.objects.get_or_create(
        payment_status=payment_status_name
    )
    shipping_status, _ = ShippingStatus.objects.get_or_create(
        shipping_status=shipping_status_name
    )
    
    # Update purchase order
    purchase_order.payment_status = payment_status
    purchase_order.shipping_status = shipping_status
    purchase_order.save()
    
    return {
        'payment_status': payment_status_name,
        'shipping_status': shipping_status_name,
        'payment_percentage': float(payment_percentage),
        'balance_remaining': float(total_amount - amount_paid)
    }


# ============================================
# PARSE DATE HELPER
# ============================================

def parse_date(date_str):
    """Parse date from DD/MM/YYYY format"""
    try:
        return datetime.strptime(date_str, '%d/%m/%Y').date()
    except:
        return None


# ============================================
# ADD NEW PAYMENT
# ============================================

@csrf_exempt
@login_required(login_url='/login/')
def api_add_payment(request):
    """
    Add new payment and automatically update PO statuses
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
    
    try:
        data = json.loads(request.body)
        
        # Extract payment data
        transaction_id = data.get('transaction_id', '').strip()
        payment_date_str = data.get('payment_date', '').strip()
        supplier_id = data.get('supplier_id', '').strip()
        supplier_name = data.get('supplier_name', '').strip()
        county_name = data.get('county', '').strip()
        town_name = data.get('town', '').strip()
        po_id = data.get('po_id', '').strip()
        bill_number = data.get('bill_number', '').strip()
        payment_mode_name = data.get('payment_mode', '').strip()
        amount_paid = Decimal(str(data.get('amount_paid', 0)))
        
        # Validation
        if not all([transaction_id, payment_date_str, supplier_name, po_id, payment_mode_name]):
            return JsonResponse({
                'success': False, 
                'message': 'All required fields must be filled'
            }, status=400)
        
        if amount_paid <= 0:
            return JsonResponse({
                'success': False, 
                'message': 'Payment amount must be greater than zero'
            }, status=400)
        
        # Check if transaction ID already exists
        if Payment.objects.filter(transaction_id=transaction_id).exists():
            return JsonResponse({
                'success': False, 
                'message': 'Transaction ID already exists'
            }, status=400)
        
        # Parse payment date
        payment_date = parse_date(payment_date_str)
        if not payment_date:
            return JsonResponse({
                'success': False, 
                'message': 'Invalid date format. Use DD/MM/YYYY'
            }, status=400)
        
        # Get foreign key objects
        try:
            supplier = Supplier.objects.get(supplier_id=supplier_id)
            purchase_order = PurchaseOrder.objects.get(po_id=po_id)
            payment_mode, _ = PaymentMode.objects.get_or_create(payment_mode=payment_mode_name)
            county = County.objects.get(county=county_name) if county_name else None
            town = Town.objects.get(town=town_name) if town_name else None
        except (Supplier.DoesNotExist, PurchaseOrder.DoesNotExist) as e:
            return JsonResponse({
                'success': False, 
                'message': f'Invalid reference: {str(e)}'
            }, status=400)
        
        # Check if payment exceeds PO balance
        balance = purchase_order.total_amount - purchase_order.amount_paid
        if amount_paid > balance:
            return JsonResponse({
                'success': False, 
                'message': f'Payment amount ({amount_paid}) exceeds remaining balance ({balance})'
            }, status=400)
        
        # Start transaction
        with transaction.atomic():
            # Create payment record
            Payment.objects.create(
                transaction_id=transaction_id,
                date=payment_date,
                supplier_id=supplier,
                supplier_name=supplier_name,
                county=county,
                town=town,
                po_id=purchase_order,
                bill_number=bill_number,
                payment_mode=payment_mode,
                amount_paid=amount_paid
            )
            
            # Update purchase order amount_paid
            purchase_order.amount_paid += amount_paid
            purchase_order.save()
            
            # Update supplier total_payments
            supplier.total_payments += amount_paid
            supplier.save()
            
            # Update PO statuses
            status_info = update_purchase_order_statuses(purchase_order)
        
        return JsonResponse({
            'success': True, 
            'message': 'Payment recorded successfully',
            'status_info': status_info
        })
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


# ============================================
# UPDATE PAYMENT
# ============================================

@csrf_exempt
@login_required(login_url='/login/')
def api_update_payment(request):
    """
    Update existing payment
    - Reverse original payment from old PO
    - Apply new payment to new PO (if changed)
    - Update both PO statuses
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
    
    try:
        data = json.loads(request.body)
        
        # Extract data
        transaction_id = data.get('transaction_id', '').strip()
        original_transaction_id = data.get('original_transaction_id', '').strip()
        original_po_id = data.get('original_po_id', '').strip()
        original_amount_paid = Decimal(str(data.get('original_amount_paid', 0)))
        
        payment_date_str = data.get('payment_date', '').strip()
        supplier_id = data.get('supplier_id', '').strip()
        supplier_name = data.get('supplier_name', '').strip()
        county_name = data.get('county', '').strip()
        town_name = data.get('town', '').strip()
        po_id = data.get('po_id', '').strip()
        bill_number = data.get('bill_number', '').strip()
        payment_mode_name = data.get('payment_mode', '').strip()
        amount_paid = Decimal(str(data.get('amount_paid', 0)))
        
        # Get existing payment
        try:
            payment = Payment.objects.get(transaction_id=original_transaction_id)
        except Payment.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Payment not found'}, status=404)
        
        # Parse date
        payment_date = parse_date(payment_date_str)
        if not payment_date:
            return JsonResponse({
                'success': False, 
                'message': 'Invalid date format. Use DD/MM/YYYY'
            }, status=400)
        
        # Get foreign key objects
        try:
            supplier = Supplier.objects.get(supplier_id=supplier_id)
            new_purchase_order = PurchaseOrder.objects.get(po_id=po_id)
            old_purchase_order = PurchaseOrder.objects.get(po_id=original_po_id)
            payment_mode, _ = PaymentMode.objects.get_or_create(payment_mode=payment_mode_name)
            county = County.objects.get(county=county_name) if county_name else None
            town = Town.objects.get(town=town_name) if town_name else None
        except (Supplier.DoesNotExist, PurchaseOrder.DoesNotExist) as e:
            return JsonResponse({
                'success': False, 
                'message': f'Invalid reference: {str(e)}'
            }, status=400)
        
        with transaction.atomic():
            # Reverse original payment from old PO
            old_purchase_order.amount_paid -= original_amount_paid
            old_purchase_order.save()
            
            # Reverse from supplier if supplier changed
            if payment.supplier_id.supplier_id != supplier_id:
                old_supplier = payment.supplier_id
                old_supplier.total_payments -= original_amount_paid
                old_supplier.save()
            
            # Apply new payment to new PO
            new_purchase_order.amount_paid += amount_paid
            new_purchase_order.save()
            
            # Update supplier total_payments
            if payment.supplier_id.supplier_id != supplier_id:
                supplier.total_payments += amount_paid
                supplier.save()
            else:
                # Same supplier, adjust by difference
                supplier.total_payments += (amount_paid - original_amount_paid)
                supplier.save()
            
            # Update payment record
            payment.transaction_id = transaction_id
            payment.date = payment_date
            payment.supplier_id = supplier
            payment.supplier_name = supplier_name
            payment.county = county
            payment.town = town
            payment.po_id = new_purchase_order
            payment.bill_number = bill_number
            payment.payment_mode = payment_mode
            payment.amount_paid = amount_paid
            payment.save()
            
            # Update statuses for both POs
            old_status_info = update_purchase_order_statuses(old_purchase_order)
            new_status_info = update_purchase_order_statuses(new_purchase_order)
        
        return JsonResponse({
            'success': True, 
            'message': 'Payment updated successfully',
            'status_info': new_status_info
        })
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


# ============================================
# DELETE PAYMENT
# ============================================

@csrf_exempt
@login_required(login_url='/login/')
def api_delete_payment(request):
    """
    Delete payment
    - Reverse payment from PO
    - Reverse from supplier
    - Update PO statuses
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
    
    try:
        data = json.loads(request.body)
        transaction_id = data.get('transaction_id', '').strip()
        
        if not transaction_id:
            return JsonResponse({'success': False, 'message': 'Transaction ID is required'}, status=400)
        
        try:
            payment = Payment.objects.get(transaction_id=transaction_id)
        except Payment.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Payment not found'}, status=404)
        
        with transaction.atomic():
            # Reverse payment from PO
            purchase_order = payment.po_id
            purchase_order.amount_paid -= payment.amount_paid
            purchase_order.save()
            
            # Reverse from supplier
            supplier = payment.supplier_id
            supplier.total_payments -= payment.amount_paid
            supplier.save()
            
            # Update PO statuses
            status_info = update_purchase_order_statuses(purchase_order)
            
            # Delete payment
            payment.delete()
        
        return JsonResponse({
            'success': True, 
            'message': 'Payment deleted successfully',
            'status_info': status_info
        })
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
@csrf_exempt
@login_required(login_url='/login/')
def api_generate_receipt_transaction_id(request):
    """Generate unique Transaction ID in format TRANSAL00001"""
    try:
        max_receipt = Receipt.objects.aggregate(Max('transaction_id'))['transaction_id__max']
        
        if max_receipt:
            num_part = int(max_receipt[7:]) + 1
        else:
            num_part = 1
        
        new_id = f"TRANSAL{num_part:05d}"
        
        while Receipt.objects.filter(transaction_id=new_id).exists():
            num_part += 1
            new_id = f"TRANSAL{num_part:05d}"
        
        return JsonResponse({'success': True, 'transaction_id': new_id})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

# =========================== TESTING THE DETAIL ID GENERATION ==================================================================================================
@csrf_exempt
@login_required(login_url='/login/')
def api_test_detail_id_generation(request):
    """
    Test endpoint to generate 10 sequential Detail IDs
    Call this to verify the generation is working correctly
    
    Usage: GET /api/purchases/test-detail-id-generation/
    """
    try:
        results = []
        
        for i in range(10):
            # Call the actual generation function
            response = api_generate_detail_id(request)
            data = response.content.decode('utf-8')
            
            import json
            result = json.loads(data)
            
            if result.get('success'):
                results.append({
                    'attempt': i + 1,
                    'detail_id': result.get('detail_id'),
                    'timestamp': time.strftime('%H:%M:%S.%f')[:-3]
                })
            
            # Small delay to ensure sequential generation
            time.sleep(0.1)
        
        return JsonResponse({
            'success': True,
            'message': 'Generated 10 sequential Detail IDs',
            'results': results
        })
    
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
@csrf_exempt
@login_required(login_url='/login/')
def api_get_receipts(request):
    """Get all receipts with full details"""
    try:
        receipts = Receipt.objects.select_related(
            'customer_id', 'county', 'town', 'so_id', 'payment_mode'
        ).all().order_by('-date')
        
        receipts_list = []
        for receipt in receipts:
            # Format date as DD/MM/YYYY
            date_str = receipt.date.strftime('%d/%m/%Y') if receipt.date else ''
            
            receipts_list.append({
                'transaction_id': receipt.transaction_id,
                'date': date_str,
                'customer_id': receipt.customer_id.customer_id if receipt.customer_id else '',
                'customer_name': receipt.customer_name,
                'county': receipt.county.county if receipt.county else '',
                'town': receipt.town.town if receipt.town else '',
                'so_id': receipt.so_id.so_id if receipt.so_id else '',
                'invoice_number': receipt.invoice_number,
                'payment_mode': receipt.payment_mode.payment_mode if receipt.payment_mode else '',
                'amount_received': float(receipt.amount_received)
            })
        
        return JsonResponse({'success': True, 'data': receipts_list})
    
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


# ============================================
# HELPER: Calculate and Update SO Statuses
# ============================================

def update_sales_order_statuses(sales_order):
    """
    Update both payment and shipping status for a sales order
    
    Receipt Status Rules:
    - 0% paid -> PENDING
    - 1-99% paid -> PARTIAL PAYMENT
    - 100% paid -> COMPLETED
    
    Shipping Status Rules:
    - 0% paid -> PENDING
    - 1-49% paid -> PROCESSING
    - 50-74% paid -> DISPATCHED
    - 75-99% paid -> IN TRANSIT
    - 100% paid -> DELIVERED
    """
    total_amount = Decimal(str(sales_order.total_amount))
    amount_received = Decimal(str(sales_order.amount_received))
    
    # Calculate payment percentage
    if total_amount > 0:
        payment_percentage = (amount_received / total_amount) * 100
    else:
        payment_percentage = 0
    
    # Determine Payment Status
    if amount_received == Decimal('0.00'):
        receipt_status_name = 'PENDING'
    elif amount_received >= total_amount:
        receipt_status_name = 'COMPLETED'
    else:
        receipt_status_name = 'PARTIAL PAYMENT'
    
    # Determine Shipping Status
    if payment_percentage == 0:
        shipping_status_name = 'PENDING'
    elif payment_percentage >= 100:
        shipping_status_name = 'DELIVERED'
    elif payment_percentage >= 75:
        shipping_status_name = 'IN TRANSIT'
    elif payment_percentage >= 50:
        shipping_status_name = 'DISPATCHED'
    else:
        shipping_status_name = 'PROCESSING'
    
    # Get or create status objects
    receipt_status, _ = ReceiptStatus.objects.get_or_create(
        receipt_status=receipt_status_name
    )
    shipping_status, _ = ShippingStatus.objects.get_or_create(
        shipping_status=shipping_status_name
    )
    
    # Update sales order
    sales_order.receipt_status = receipt_status
    sales_order.shipping_status = shipping_status
    sales_order.save()
    
    return {
        'receipt_status': receipt_status_name,
        'shipping_status': shipping_status_name,
        'payment_percentage': float(payment_percentage),
        'balance_remaining': float(total_amount - amount_received)
    }

# ============================================
# ADD NEW RECEIPT
# ============================================

@csrf_exempt
@login_required(login_url='/login/')
def api_add_receipt(request):
    """
    Add new receipt and automatically update SO statuses
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
    
    try:
        data = json.loads(request.body)
        
        # Extract receipt data
        transaction_id = data.get('transaction_id', '').strip()
        receipt_date_str = data.get('receipt_date', '').strip()
        customer_id = data.get('customer_id', '').strip()
        customer_name = data.get('customer_name', '').strip()
        county_name = data.get('county', '').strip()
        town_name = data.get('town', '').strip()
        so_id = data.get('so_id', '').strip()
        invoice_number = data.get('invoice_number', '').strip()
        payment_mode_name = data.get('payment_mode', '').strip()
        amount_received = Decimal(str(data.get('amount_received', 0)))
        
        # Validation
        if not all([transaction_id, receipt_date_str, customer_name, so_id, payment_mode_name]):
            return JsonResponse({
                'success': False, 
                'message': 'All required fields must be filled'
            }, status=400)
        
        if amount_received <= 0:
            return JsonResponse({
                'success': False, 
                'message': 'Payment amount must be greater than zero'
            }, status=400)
        
        # Check if transaction ID already exists
        if Receipt.objects.filter(transaction_id=transaction_id).exists():
            return JsonResponse({
                'success': False, 
                'message': 'Transaction ID already exists'
            }, status=400)
        
        # Parse receipt date
        receipt_date = parse_date(receipt_date_str)
        if not receipt_date:
            return JsonResponse({
                'success': False, 
                'message': 'Invalid date format. Use DD/MM/YYYY'
            }, status=400)
        
        # Get foreign key objects
        try:
            customer = Customer.objects.get(customer_id=customer_id)
            sales_order = SalesOrder.objects.get(so_id=so_id)
            payment_mode, _ = PaymentMode.objects.get_or_create(payment_mode=payment_mode_name)
            county = County.objects.get(county=county_name) if county_name else None
            town = Town.objects.get(town=town_name) if town_name else None
        except (Customer.DoesNotExist, SalesOrder.DoesNotExist) as e:
            return JsonResponse({
                'success': False, 
                'message': f'Invalid reference: {str(e)}'
            }, status=400)
        
        # Check if payment exceeds SO balance
        balance = sales_order.total_amount - sales_order.amount_received
        if amount_received > balance:
            return JsonResponse({
                'success': False, 
                'message': f'Payment amount ({amount_received}) exceeds remaining balance ({balance})'
            }, status=400)
        
        # Start transaction
        with transaction.atomic():
            # Create receipt record
            Receipt.objects.create(
                transaction_id=transaction_id,
                date=receipt_date,
                customer_id=customer,
                customer_name=customer_name,
                county=county,
                town=town,
                so_id=sales_order,
                invoice_number=invoice_number,
                payment_mode=payment_mode,
                amount_received=amount_received
            )
            
            # Update sales order amount_received
            sales_order.amount_received += amount_received
            sales_order.save()
            
            # Update customer total_payments
            customer.total_payments += amount_received
            customer.save()
            
            # *** UPDATE SO STATUSES ***
            status_info = update_sales_order_statuses(sales_order)
        
        return JsonResponse({
            'success': True, 
            'message': 'Receipt recorded successfully',
            'status_info': status_info
        })
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

@csrf_exempt
@login_required(login_url='/login/')
def api_delete_receipt(request):
    """
    Delete receipt
    - Reverse receipt from SO
    - Reverse from customer
    - Update SO statuses
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
    
    try:
        data = json.loads(request.body)
        transaction_id = data.get('transaction_id', '').strip()
        
        if not transaction_id:
            return JsonResponse({'success': False, 'message': 'Transaction ID is required'}, status=400)
        
        try:
            receipt = Receipt.objects.get(transaction_id=transaction_id)
        except Receipt.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Receipt not found'}, status=404)
        
        with transaction.atomic():
            # Get the sales order
            sales_order = receipt.so_id
            
            # Reverse receipt from SO
            sales_order.amount_received -= receipt.amount_received
            sales_order.save()
            
            # Reverse from customer
            customer = receipt.customer_id
            customer.total_payments -= receipt.amount_received
            customer.save()
            
            # Update SO statuses
            status_info = update_sales_order_statuses(sales_order)
            
            # Delete receipt
            receipt.delete()
        
        return JsonResponse({
            'success': True, 
            'message': 'Receipt deleted successfully',
            'status_info': status_info
        })
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        print(f"❌ Error deleting receipt: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
# ============================================
# UPDATE RECEIPT
# ============================================

@csrf_exempt
@login_required(login_url='/login/')
def api_update_receipt(request):
    """
    Update existing payment
    - Reverse original payment from old PO
    - Apply new payment to new PO (if changed)
    - Update both PO statuses
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
    
    try:
        data = json.loads(request.body)
        
        # Extract data
        transaction_id = data.get('transaction_id', '').strip()
        original_transaction_id = data.get('original_transaction_id', '').strip()
        original_so_id = data.get('original_so_id', '').strip()
        original_amount_received = Decimal(str(data.get('original_amount_received ', 0)))
        
        receipt_date_str = data.get('receipt_date', '').strip()
        customer_id = data.get('customer_id', '').strip()
        customer_name = data.get('customer_name', '').strip()
        county_name = data.get('county', '').strip()
        town_name = data.get('town', '').strip()
        so_id = data.get('so_id', '').strip()
        invoice_number = data.get('invoice_number', '').strip()
        payment_mode_name = data.get('payment_mode', '').strip()
        amount_received = Decimal(str(data.get('amount_received', 0)))
        
        # Get existing receipt
        try:
            receipt = Receipt.objects.get(transaction_id=original_transaction_id)
        except Receipt.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Receipt not found'}, status=404)
        
        # Parse date
        receipt_date = parse_date(receipt_date_str)
        if not receipt_date:
            return JsonResponse({
                'success': False, 
                'message': 'Invalid date format. Use DD/MM/YYYY'
            }, status=400)
        
        # Get foreign key objects
        try:
            customer = Customer.objects.get(customer_id=customer_id)
            new_sales_order = SalesOrder.objects.get(so_id=so_id)
            old_sales_order = SalesOrder.objects.get(so_id=original_so_id)
            payment_mode, _ = PaymentMode.objects.get_or_create(payment_mode=payment_mode_name)
            county = County.objects.get(county=county_name) if county_name else None
            town = Town.objects.get(town=town_name) if town_name else None
        except (Customer.DoesNotExist, SalesOrder.DoesNotExist) as e:
            return JsonResponse({
                'success': False, 
                'message': f'Invalid reference: {str(e)}'
            }, status=400)
        
        with transaction.atomic():
            # Reverse original receipt from old SO
            old_sales_order.amount_received -= original_amount_received
            old_sales_order.save()
            
            # Reverse from customer if supplier changed
            if receipt.customer_id.customer_id != customer_id:
                old_customer = receipt.customer_id
                old_customer.total_payments -= original_amount_received
                old_customer.save()
            
            # Apply new payment to new PO
            new_sales_order.amount_received += amount_received
            new_sales_order.save()
            
            # Update supplier total_payments
            if customer.customer_id.customer_id != customer_id:
                customer.total_payments += amount_received
                customer.save()
            else:
                # Same customer, adjust by difference
                customer.total_payments += (amount_received - original_amount_received)
                customer.save()
            
            # Update payment record
            receipt.transaction_id = transaction_id
            receipt.date = receipt_date
            receipt.customer_id = customer
            receipt.customer_name = customer_name
            receipt.county = county
            receipt.town = town
            receipt.so_id = new_sales_order
            receipt.invoice_number = invoice_number
            receipt.payment_mode = payment_mode
            receipt.amount_received = amount_received
            receipt.save()
            
            # Update statuses for both SOs
            old_status_info = update_sales_order_statuses(old_sales_order)
            new_status_info = update_sales_order_statuses(new_sales_order)
        
        return JsonResponse({
            'success': True, 
            'message': 'Payment updated successfully',
            'status_info': new_status_info
        })
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
    
# =============================== DASHBOARD VIEWS ==============================================================================================================================================
# ============================================
# NEW ENDPOINT: Get All Sales Details
# ============================================
@csrf_exempt
@login_required(login_url='/login/')
def api_get_all_sales_details(request):
    """
    Get all sales details efficiently in a single call
    Used for: Top Selling Item, Sales by Category charts
    """
    try:
        sales_details = SalesDetail.objects.select_related(
            'so_id', 'customer_id', 'item_id'
        ).all()
        
        details_list = []
        for detail in sales_details:
            details_list.append({
                'detail_id': detail.detail_id,
                'item_name': detail.item_name,
                'item_type': detail.item_type,
                'item_category': detail.item_category,
                'item_subcategory': detail.item_subcategory,
                'quantity_sold': detail.quantity_sold,
                'total_sales_price': float(detail.total_sales_price)
            })
        
        return JsonResponse({
            'success': True,
            'data': details_list,
            'count': len(details_list)
        })
    
    except Exception as e:
        print(f"❌ Error getting sales details: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


# ============================================
# NEW ENDPOINT: Get All Purchase Details
# ============================================
@csrf_exempt
@login_required(login_url='/login/')
def api_get_all_purchase_details(request):
    """
    Get all purchase details efficiently in a single call
    Used for: Purchases by Category chart
    """
    try:
        purchase_details = PurchaseDetail.objects.select_related(
            'po_id', 'supplier_id', 'item_id'
        ).all()
        
        details_list = []
        for detail in purchase_details:
            details_list.append({
                'detail_id': detail.detail_id,
                'item_name': detail.item_name,
                'item_type': detail.item_type,
                'item_category': detail.item_category,
                'item_subcategory': detail.item_subcategory,
                'quantity_purchased': detail.quantity_purchased,
                'total_purchase_price': float(detail.total_purchase_price)
            })
        
        return JsonResponse({
            'success': True,
            'data': details_list,
            'count': len(details_list)
        })
    
    except Exception as e:
        print(f"❌ Error getting purchase details: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)

# ========================= SALES FORECASTING VIEWS ============================================================================================================================

# Django backend endpoint
def calculate_forecast_ensemble(historical_data, periods):
    forecasts = {
        'linear': linear_regression_forecast(),
        'ema': exponential_smoothing_forecast(),
        'arima': arima_forecast(),
    }
    # Weighted average of all models
    ensemble = weighted_average(forecasts, weights={'linear': 0.3, 'ema': 0.4, 'arima': 0.3})
    return ensemble

def ml_forecast(sales_history, external_features):
    # Features: previous 3 months, day of week, is_holiday, etc.
    X = prepare_features(sales_history, external_features)
    y = sales_history['revenue']
    
    model = RandomForestRegressor(n_estimators=100)
    model.fit(X, y)
    predictions = model.predict(future_features)
    return predictions

def calculate_accuracy_metrics(forecasts, actuals):
    mape = np.mean(np.abs((actuals - forecasts) / actuals)) * 100
    rmse = np.sqrt(np.mean((actuals - forecasts) ** 2))
    return {'mape': mape, 'rmse': rmse, 'accuracy': 100 - mape}

def calculate_eoq(annual_demand, ordering_cost, holding_cost):
    """Economic Order Quantity"""
    eoq = math.sqrt((2 * annual_demand * ordering_cost) / holding_cost)
    return eoq

def calculate_safety_stock(forecast_std, service_level=0.95):
    """Safety stock for 95% service level"""
    z_score = 1.65  # 95% service level
    safety_stock = z_score * forecast_std
    return safety_stock

def forecast_profitability(sales_forecast, inventory_data):
    revenue = sales_forecast['total_sales']
    cogs = sum(item.quantity * item.purchase_price for item in inventory_data)
    gross_profit = revenue - cogs
    gross_margin = (gross_profit / revenue) * 100
    
    return {
        'revenue': revenue,
        'cogs': cogs,
        'gross_profit': gross_profit,
        'gross_margin': gross_margin
    }
    



# ============================================ REPORTS VIEWS ==================================================================================================================

# HELPER FUNCTIONS
# ============================================

def parse_date(date_str):
    """Parse date from DD/MM/YYYY or YYYY-MM-DD format"""
    if not date_str:
        return None
    
    # Try DD/MM/YYYY format first
    try:
        return datetime.strptime(date_str, '%d/%m/%Y').date()
    except:
        pass
    
    # Try YYYY-MM-DD format
    try:
        return datetime.strptime(date_str, '%Y-%m-%d').date()
    except:
        return None


def format_currency(amount):
    """Format amount as currency string"""
    if amount is None:
        return "0.00"
    return f"{float(amount):,.2f}"


def get_date_range(range_type):
    """
    Get start and end dates for quick report ranges
    Returns tuple: (start_date, end_date)
    """
    from datetime import date
    today = date.today()
    
    if range_type == 'today':
        return today, today
    
    elif range_type == 'week':
        # Get start of current week (Monday)
        start = today - timedelta(days=today.weekday())
        return start, today
    
    elif range_type == 'month':
        # Get start of current month
        start = date(today.year, today.month, 1)
        return start, today
    
    elif range_type == 'year':
        # Get start of current year
        start = date(today.year, 1, 1)
        return start, today
    
    return None, None

# ============================================
# 1. SALES SUMMARY REPORT
# ============================================
@csrf_exempt
@login_required(login_url='/login/')
def api_generate_sales_summary(request):
    """
    Generate Sales Summary Report
    Returns: Aggregated sales data with charts
    """
    try:
        # Get filter parameters
        start_date_str = request.GET.get('start_date', '')
        end_date_str = request.GET.get('end_date', '')
        category = request.GET.get('category', '')
        
        # Parse dates
        start_date = parse_date(start_date_str) if start_date_str else None
        end_date = parse_date(end_date_str) if end_date_str else None
        
        # Build query
        query = SalesOrder.objects.all()
        
        if start_date:
            query = query.filter(date__gte=start_date)
        if end_date:
            query = query.filter(date__lte=end_date)
        
        # Get sales orders
        sales_orders = query.select_related('customer_id', 'county', 'town')
        
        # Calculate KPIs
        total_sales = sales_orders.aggregate(
            total=Sum('total_amount')
        )['total'] or Decimal('0.00')
        
        total_received = sales_orders.aggregate(
            received=Sum('amount_received')
        )['received'] or Decimal('0.00')
        
        total_orders = sales_orders.count()
        
        outstanding = total_sales - total_received
        
        # Sales by category - FIXED: Calculate from detail records
        category_query = SalesDetail.objects.filter(
            so_id__in=sales_orders
        )
        
        if category:
            category_query = category_query.filter(item_category=category)
        
        # ✅ FIX: Calculate total manually instead of using @property
        sales_by_category_raw = category_query.values('item_category').annotate(
            quantity=Sum('quantity_sold'),
            subtotal=Sum(F('quantity_sold') * F('unit_price')),
            tax=Sum(F('quantity_sold') * F('unit_price') * F('tax_rate') / 100)
        ).order_by('-subtotal')[:10]
        
        # Calculate final totals with shipping
        sales_by_category = []
        for item in sales_by_category_raw:
            subtotal = item['subtotal'] or Decimal('0.00')
            tax = item['tax'] or Decimal('0.00')
            price_with_tax = subtotal + tax
            shipping = price_with_tax * Decimal('0.02')  # 2% shipping
            total = price_with_tax + shipping
            
            sales_by_category.append({
                'category': item['item_category'],
                'total': float(total)
            })
        
        # Sales trend (daily)
        daily_sales = sales_orders.annotate(
            sale_date=TruncDate('date')
        ).values('sale_date').annotate(
            total=Sum('total_amount')
        ).order_by('sale_date')
        
        # Top customers
        top_customers = sales_orders.values(
            'customer_name'
        ).annotate(
            total=Sum('total_amount'),
            orders=Count('so_id')
        ).order_by('-total')[:5]
        
        # Sales by location
        sales_by_location = sales_orders.values(
            'county__county'
        ).annotate(
            total=Sum('total_amount')
        ).order_by('-total')[:10]
        
        # Format data for response
        report_data = {
            'kpis': {
                'total_sales': float(total_sales),
                'total_received': float(total_received),
                'outstanding': float(outstanding),
                'total_orders': total_orders,
                'avg_order_value': float(total_sales / total_orders) if total_orders > 0 else 0
            },
            'sales_by_category': sales_by_category,
            'daily_sales': [
                {
                    'date': item['sale_date'].strftime('%Y-%m-%d'),
                    'total': float(item['total'])
                }
                for item in daily_sales
            ],
            'top_customers': [
                {
                    'name': item['customer_name'],
                    'total': float(item['total']),
                    'orders': item['orders']
                }
                for item in top_customers
            ],
            'sales_by_location': [
                {
                    'location': item['county__county'] or 'Unknown',
                    'total': float(item['total'])
                }
                for item in sales_by_location
            ],
            'period': {
                'start': start_date.strftime('%d/%m/%Y') if start_date else '',
                'end': end_date.strftime('%d/%m/%Y') if end_date else ''
            }
        }
        
        return JsonResponse({
            'success': True,
            'data': report_data,
            'report_type': 'Sales Summary Report'
        })
    
    except Exception as e:
        print(f"Error generating sales summary: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)

# ============================================
# 2. INVENTORY STATUS REPORT
# ============================================

@csrf_exempt
@login_required(login_url='/login/')
def api_generate_inventory_status(request):
    """
    Generate Inventory Status Report
    Returns: Current inventory levels and valuations
    """
    try:
        category = request.GET.get('category', '')
        
        # Build query
        query = Inventory.objects.select_related(
            'item_type', 'item_category', 'item_subcategory'
        )
        
        if category:
            query = query.filter(item_category__item_category=category)
        
        inventory_items = query.all()
        
        # Calculate totals
        total_cost_value = Decimal('0.00')
        total_sale_value = Decimal('0.00')
        total_items = 0
        reorder_items = 0
        
        items_list = []
        low_stock_items = []
        
        for item in inventory_items:
            remaining_qty = item.quantity_purchased - item.quantity_sold
            cost_value = remaining_qty * item.purchase_price
            sale_value = remaining_qty * item.sale_price
            
            total_cost_value += cost_value
            total_sale_value += sale_value
            total_items += 1
            
            if item.reorder_required == 'YES':
                reorder_items += 1
                low_stock_items.append({
                    'item_id': item.item_id,
                    'name': item.item_name,
                    'category': item.item_category.item_category,
                    'remaining_qty': remaining_qty,
                    'reorder_level': item.reorder_level
                })
            
            items_list.append({
                'item_id': item.item_id,
                'name': item.item_name,
                'type': item.item_type.item_type,
                'category': item.item_category.item_category,
                'subcategory': item.item_subcategory.item_subcategory,
                'purchased_qty': item.quantity_purchased,
                'sold_qty': item.quantity_sold,
                'remaining_qty': remaining_qty,
                'purchase_price': float(item.purchase_price),
                'sale_price': float(item.sale_price),
                'cost_value': float(cost_value),
                'sale_value': float(sale_value),
                'reorder_required': item.reorder_required
            })
        
        potential_profit = total_sale_value - total_cost_value
        
        # Inventory by category
        by_category = {}
        for item_data in items_list:
            cat = item_data['category']
            if cat not in by_category:
                by_category[cat] = {
                    'cost_value': 0,
                    'sale_value': 0,
                    'items': 0
                }
            by_category[cat]['cost_value'] += item_data['cost_value']
            by_category[cat]['sale_value'] += item_data['sale_value']
            by_category[cat]['items'] += 1
        
        category_breakdown = [
            {
                'category': cat,
                'cost_value': data['cost_value'],
                'sale_value': data['sale_value'],
                'items': data['items']
            }
            for cat, data in by_category.items()
        ]
        
        report_data = {
            'kpis': {
                'total_cost_value': float(total_cost_value),
                'total_sale_value': float(total_sale_value),
                'potential_profit': float(potential_profit),
                'total_items': total_items,
                'reorder_items': reorder_items
            },
            'items': items_list,
            'low_stock': low_stock_items,
            'category_breakdown': category_breakdown
        }
        
        return JsonResponse({
            'success': True,
            'data': report_data,
            'report_type': 'Inventory Status Report'
        })
    
    except Exception as e:
        print(f"Error generating inventory status: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


# ============================================
# 3. PROFIT & LOSS REPORT
# ============================================
@csrf_exempt
@login_required(login_url='/login/')
def api_generate_profit_loss(request):
    """
    Generate Profit & Loss Report
    Returns: Revenue, expenses, and profit calculations
    """
    try:
        start_date_str = request.GET.get('start_date', '')
        end_date_str = request.GET.get('end_date', '')
        
        start_date = parse_date(start_date_str) if start_date_str else None
        end_date = parse_date(end_date_str) if end_date_str else None
        
        # Sales Revenue
        sales_query = SalesOrder.objects.all()
        if start_date:
            sales_query = sales_query.filter(date__gte=start_date)
        if end_date:
            sales_query = sales_query.filter(date__lte=end_date)
        
        total_revenue = sales_query.aggregate(
            total=Sum('total_amount')
        )['total'] or Decimal('0.00')
        
        # Get sales details for COGS calculation
        sales_details = SalesDetail.objects.filter(so_id__in=sales_query)
        
        # Calculate COGS based on items sold and their purchase prices
        cogs = Decimal('0.00')
        for detail in sales_details:
            item = detail.item_id
            if item:
                cogs += detail.quantity_sold * item.purchase_price
        
        # Gross Profit
        gross_profit = total_revenue - cogs
        gross_margin = (gross_profit / total_revenue * 100) if total_revenue > 0 else 0
        
        # ✅ FIX: Calculate shipping expenses manually instead of aggregating @property
        shipping_expense = Decimal('0.00')
        for detail in sales_details:
            # Calculate shipping for each detail: 2% of (price + tax)
            subtotal = detail.quantity_sold * detail.unit_price
            tax = subtotal * detail.tax_rate / 100
            price_with_tax = subtotal + tax
            shipping = price_with_tax * Decimal('0.02')  # 2% shipping
            shipping_expense += shipping
            
        """     ALTERNATIVE FUNCTION/LOGIC FOR CALCULATING OVERALL SHIPPING FEES
        shipping_data = sales_details.aggregate(
            shipping_total=Sum(
                ExpressionWrapper(
                    (F('quantity_sold') * F('unit_price') * (1 + F('tax_rate') / 100)) * Decimal('0.02'),
                    output_field=DecimalField(max_digits=12, decimal_places=2)
                )
            )
        )
        shipping_expense = shipping_data['shipping_total'] or Decimal('0.00') 
        """   
        
        # Purchases in period
        purchase_query = PurchaseOrder.objects.all()
        if start_date:
            purchase_query = purchase_query.filter(date__gte=start_date)
        if end_date:
            purchase_query = purchase_query.filter(date__lte=end_date)
        
        total_purchases = purchase_query.aggregate(
            total=Sum('total_amount')
        )['total'] or Decimal('0.00')
        
        # Net Profit
        total_expenses = shipping_expense
        net_profit = gross_profit - total_expenses
        net_margin = (net_profit / total_revenue * 100) if total_revenue > 0 else 0
        
        # Monthly breakdown
        monthly_data = sales_query.annotate(
            month=TruncMonth('date')
        ).values('month').annotate(
            revenue=Sum('total_amount')
        ).order_by('month')
        
        report_data = {
            'kpis': {
                'total_revenue': float(total_revenue),
                'cogs': float(cogs),
                'gross_profit': float(gross_profit),
                'gross_margin': float(gross_margin),
                'operating_expenses': float(total_expenses),
                'net_profit': float(net_profit),
                'net_margin': float(net_margin)
            },
            'breakdown': {
                'revenue': float(total_revenue),
                'cogs': float(cogs),
                'shipping_expense': float(shipping_expense),
                'gross_profit': float(gross_profit),
                'net_profit': float(net_profit)
            },
            'monthly_trend': [
                {
                    'month': item['month'].strftime('%b %Y'),
                    'revenue': float(item['revenue'])
                }
                for item in monthly_data
            ],
            'period': {
                'start': start_date.strftime('%d/%m/%Y') if start_date else '',
                'end': end_date.strftime('%d/%m/%Y') if end_date else ''
            }
        }
        
        return JsonResponse({
            'success': True,
            'data': report_data,
            'report_type': 'Profit & Loss Report'
        })
    
    except Exception as e:
        print(f"❌ Error generating P&L report: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)
        
# ============================================
# 4. PURCHASE SUMMARY REPORT
# ============================================
@csrf_exempt
@login_required(login_url='/login/')
def api_generate_purchase_summary(request):
    """Generate Purchase Summary Report"""
    try:
        start_date_str = request.GET.get('start_date', '')
        end_date_str = request.GET.get('end_date', '')
        
        start_date = parse_date(start_date_str) if start_date_str else None
        end_date = parse_date(end_date_str) if end_date_str else None
        
        query = PurchaseOrder.objects.all()
        
        if start_date:
            query = query.filter(date__gte=start_date)
        if end_date:
            query = query.filter(date__lte=end_date)
        
        purchase_orders = query.select_related('supplier_id', 'county', 'town')
        
        # KPIs
        total_purchases = purchase_orders.aggregate(
            total=Sum('total_amount')
        )['total'] or Decimal('0.00')
        
        total_paid = purchase_orders.aggregate(
            paid=Sum('amount_paid')
        )['paid'] or Decimal('0.00')
        
        outstanding = total_purchases - total_paid
        total_orders = purchase_orders.count()
        
        # Purchases by category - FIXED
        purchase_details = PurchaseDetail.objects.filter(
            po_id__in=purchase_orders
        )
        
        # ✅ FIX: Calculate total manually
        purchases_by_category_raw = purchase_details.values('item_category').annotate(
            quantity=Sum('quantity_purchased'),
            subtotal=Sum(F('quantity_purchased') * F('unit_cost')),
            tax=Sum(F('quantity_purchased') * F('unit_cost') * F('tax_rate') / 100)
        ).order_by('-subtotal')[:10]
        
        # Calculate final totals with shipping
        purchases_by_category = []
        for item in purchases_by_category_raw:
            subtotal = item['subtotal'] or Decimal('0.00')
            tax = item['tax'] or Decimal('0.00')
            cost_with_tax = subtotal + tax
            shipping = cost_with_tax * Decimal('0.01')  # 1% shipping
            total = cost_with_tax + shipping
            
            purchases_by_category.append({
                'category': item['item_category'],
                'total': float(total)
            })
        
        # Top suppliers
        top_suppliers = purchase_orders.values(
            'supplier_name'
        ).annotate(
            total=Sum('total_amount'),
            orders=Count('po_id')
        ).order_by('-total')[:5]
        
        # Daily purchases
        daily_purchases = purchase_orders.annotate(
            purchase_date=TruncDate('date')
        ).values('purchase_date').annotate(
            total=Sum('total_amount')
        ).order_by('purchase_date')
        
        report_data = {
            'kpis': {
                'total_purchases': float(total_purchases),
                'total_paid': float(total_paid),
                'outstanding': float(outstanding),
                'total_orders': total_orders,
                'avg_order_value': float(total_purchases / total_orders) if total_orders > 0 else 0
            },
            'purchases_by_category': purchases_by_category,
            'top_suppliers': [
                {
                    'name': item['supplier_name'],
                    'total': float(item['total']),
                    'orders': item['orders']
                }
                for item in top_suppliers
            ],
            'daily_purchases': [
                {
                    'date': item['purchase_date'].strftime('%Y-%m-%d'),
                    'total': float(item['total'])
                }
                for item in daily_purchases
            ],
            'period': {
                'start': start_date.strftime('%d/%m/%Y') if start_date else '',
                'end': end_date.strftime('%d/%m/%Y') if end_date else ''
            }
        }
        
        return JsonResponse({
            'success': True,
            'data': report_data,
            'report_type': 'Purchase Summary Report'
        })
    
    except Exception as e:
        print(f"❌ Error generating purchase summary: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
# ============================================
# 5. OUTSTANDING BALANCES REPORT
# ============================================

@csrf_exempt
@login_required(login_url='/login/')
def api_generate_outstanding_balances(request):
    """Generate Outstanding Balances Report"""
    try:
        # Customer balances
        customers = Customer.objects.all()
        customer_balances = []
        total_receivable = Decimal('0.00')
        
        for customer in customers:
            balance = customer.total_sales - customer.total_payments
            if balance > 0:
                total_receivable += balance
                customer_balances.append({
                    'id': customer.customer_id,
                    'name': customer.customer_name,
                    'total_sales': float(customer.total_sales),
                    'total_payments': float(customer.total_payments),
                    'balance': float(balance)
                })
        
        # Supplier balances
        suppliers = Supplier.objects.all()
        supplier_balances = []
        total_payable = Decimal('0.00')
        
        for supplier in suppliers:
            balance = supplier.total_purchases - supplier.total_payments
            if balance > 0:
                total_payable += balance
                supplier_balances.append({
                    'id': supplier.supplier_id,
                    'name': supplier.supplier_name,
                    'total_purchases': float(supplier.total_purchases),
                    'total_payments': float(supplier.total_payments),
                    'balance': float(balance)
                })
        
        report_data = {
            'kpis': {
                'total_receivable': float(total_receivable),
                'total_payable': float(total_payable),
                'net_position': float(total_receivable - total_payable)
            },
            'customer_balances': sorted(customer_balances, key=lambda x: x['balance'], reverse=True),
            'supplier_balances': sorted(supplier_balances, key=lambda x: x['balance'], reverse=True)
        }
        
        return JsonResponse({
            'success': True,
            'data': report_data,
            'report_type': 'Outstanding Balances Report'
        })
    
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
# ============================================
# CUSTOMER ANALYSIS REPORT
# ============================================
@csrf_exempt
@login_required(login_url='/login/')
def api_generate_customer_analysis(request):
    """
    Generate Customer Analysis Report
    Returns: Customer purchasing behavior, value segmentation, and trends
    """
    try:
        start_date_str = request.GET.get('start_date', '')
        end_date_str = request.GET.get('end_date', '')
        county = request.GET.get('county', '')
        
        start_date = parse_date(start_date_str) if start_date_str else None
        end_date = parse_date(end_date_str) if end_date_str else None
        
        # Build query
        sales_query = SalesOrder.objects.all()
        
        if start_date:
            sales_query = sales_query.filter(date__gte=start_date)
        if end_date:
            sales_query = sales_query.filter(date__lte=end_date)
        
        # Filter by county if specified
        if county:
            sales_query = sales_query.filter(county__county=county)
        
        # Get all customers
        all_customers = Customer.objects.all()
        if county:
            all_customers = all_customers.filter(county__county=county)
        
        # Customer metrics
        customer_data = []
        total_customers = all_customers.count()
        active_customers = 0
        total_customer_value = Decimal('0.00')
        total_customer_orders = 0
        
        for customer in all_customers:
            # Get sales for this customer in the period
            customer_sales = sales_query.filter(customer_id=customer)
            
            sales_total = customer_sales.aggregate(
                total=Sum('total_amount')
            )['total'] or Decimal('0.00')
            
            orders_count = customer_sales.count()
            
            # Calculate payment ratio
            payments_total = customer.total_payments
            payment_ratio = (payments_total / customer.total_sales * 100) if customer.total_sales > 0 else 0
            
            # Outstanding balance
            outstanding = customer.total_sales - customer.total_payments
            
            if sales_total > 0:
                active_customers += 1
                total_customer_value += sales_total
                total_customer_orders += orders_count
                
                # Get first and last order dates
                first_order = customer_sales.order_by('date').first()
                last_order = customer_sales.order_by('-date').first()
                
                # Calculate average days between orders
                if orders_count > 1 and first_order and last_order:
                    days_span = (last_order.date - first_order.date).days
                    avg_days_between = days_span / (orders_count - 1) if orders_count > 1 else 0
                else:
                    avg_days_between = 0
                
                customer_data.append({
                    'customer_id': customer.customer_id,
                    'customer_name': customer.customer_name,
                    'phone': customer.phone_number or '',
                    'email': customer.email or '',
                    'county': customer.county.county if customer.county else '',
                    'town': customer.town.town if customer.town else '',
                    'total_sales': float(sales_total),
                    'total_orders': orders_count,
                    'avg_order_value': float(sales_total / orders_count) if orders_count > 0 else 0,
                    'lifetime_value': float(customer.total_sales),
                    'total_payments': float(payments_total),
                    'outstanding_balance': float(outstanding),
                    'payment_ratio': float(payment_ratio),
                    'first_order_date': first_order.date.strftime('%d/%m/%Y') if first_order else '',
                    'last_order_date': last_order.date.strftime('%d/%m/%Y') if last_order else '',
                    'avg_days_between_orders': int(avg_days_between)
                })
        
        # Sort by total sales
        customer_data.sort(key=lambda x: x['total_sales'], reverse=True)
        
        # Customer segmentation (RFM-like analysis)
        # Recency, Frequency, Monetary
        if customer_data:
            # Top 20% by value = VIP customers
            vip_threshold_index = max(1, int(len(customer_data) * 0.2))
            vip_customers = customer_data[:vip_threshold_index]
            
            # Next 30% = High-value customers
            high_value_threshold_index = max(vip_threshold_index + 1, int(len(customer_data) * 0.5))
            high_value_customers = customer_data[vip_threshold_index:high_value_threshold_index]
            
            # Remaining = Regular customers
            regular_customers = customer_data[high_value_threshold_index:]
            
            segmentation = {
                'vip': {
                    'count': len(vip_customers),
                    'total_value': sum(c['total_sales'] for c in vip_customers),
                    'percentage': (len(vip_customers) / len(customer_data) * 100) if customer_data else 0
                },
                'high_value': {
                    'count': len(high_value_customers),
                    'total_value': sum(c['total_sales'] for c in high_value_customers),
                    'percentage': (len(high_value_customers) / len(customer_data) * 100) if customer_data else 0
                },
                'regular': {
                    'count': len(regular_customers),
                    'total_value': sum(c['total_sales'] for c in regular_customers),
                    'percentage': (len(regular_customers) / len(customer_data) * 100) if customer_data else 0
                }
            }
        else:
            segmentation = {
                'vip': {'count': 0, 'total_value': 0, 'percentage': 0},
                'high_value': {'count': 0, 'total_value': 0, 'percentage': 0},
                'regular': {'count': 0, 'total_value': 0, 'percentage': 0}
            }
        
        # Geographic distribution
        geographic_data = all_customers.values('county__county').annotate(
            customer_count=Count('customer_id'),
            total_sales=Sum('total_sales')
        ).order_by('-total_sales')[:10]
        
        # Payment behavior analysis
        payment_analysis = {
            'excellent': 0,  # >90% payment ratio
            'good': 0,       # 70-90%
            'fair': 0,       # 50-70%
            'poor': 0        # <50%
        }
        
        for customer in customer_data:
            ratio = customer['payment_ratio']
            if ratio >= 90:
                payment_analysis['excellent'] += 1
            elif ratio >= 70:
                payment_analysis['good'] += 1
            elif ratio >= 50:
                payment_analysis['fair'] += 1
            else:
                payment_analysis['poor'] += 1
        
        # ✅ FIXED: Customer acquisition trend (new customers over time)
        # Calculate based on first order date since we don't have a customer created_date
        customer_first_orders = []
        for customer in all_customers:
            first_order = SalesOrder.objects.filter(customer_id=customer).order_by('date').first()
            if first_order:
                customer_first_orders.append({
                    'customer_id': customer.customer_id,
                    'first_order_date': first_order.date
                })
        
        # Group by month
        from collections import defaultdict
        acquisition_by_month = defaultdict(int)
        
        for item in customer_first_orders:
            month_key = item['first_order_date'].strftime('%Y-%m')
            acquisition_by_month[month_key] += 1
        
        # Sort and take last 12 months
        acquisition_trend = sorted(
            [{'month': k, 'new_customers': v} for k, v in acquisition_by_month.items()],
            key=lambda x: x['month']
        )[-12:]
        
        # Format response
        report_data = {
            'kpis': {
                'total_customers': total_customers,
                'active_customers': active_customers,
                'inactive_customers': total_customers - active_customers,
                'total_revenue': float(total_customer_value),
                'total_orders': total_customer_orders,
                'avg_customer_value': float(total_customer_value / active_customers) if active_customers > 0 else 0,
                'avg_orders_per_customer': float(total_customer_orders / active_customers) if active_customers > 0 else 0
            },
            'top_customers': customer_data[:10],  # Top 10 customers
            'all_customers': customer_data,  # All customer data
            'segmentation': segmentation,
            'geographic_distribution': [
                {
                    'county': item['county__county'] or 'Unknown',
                    'customer_count': item['customer_count'],
                    'total_sales': float(item['total_sales'] or 0)
                }
                for item in geographic_data
            ],
            'payment_behavior': payment_analysis,
            'acquisition_trend': acquisition_trend,  # Added acquisition trend data
            'period': {
                'start': start_date.strftime('%d/%m/%Y') if start_date else '',
                'end': end_date.strftime('%d/%m/%Y') if end_date else ''
            }
        }
        
        return JsonResponse({
            'success': True,
            'data': report_data,
            'report_type': 'Customer Analysis Report'
        })
    
    except Exception as e:
        print(f"❌ Error generating customer analysis: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)
            
# ============================================
# SUPPLIER ANALYSIS REPORT
# ============================================

@csrf_exempt
@login_required(login_url='/login/')
def api_generate_supplier_analysis(request):
    """
    Generate Supplier Analysis Report
    Returns: Supplier performance, reliability, and cost analysis
    """
    try:
        start_date_str = request.GET.get('start_date', '')
        end_date_str = request.GET.get('end_date', '')
        county = request.GET.get('county', '')
        
        start_date = parse_date(start_date_str) if start_date_str else None
        end_date = parse_date(end_date_str) if end_date_str else None
        
        # Build query
        purchase_query = PurchaseOrder.objects.all()
        
        if start_date:
            purchase_query = purchase_query.filter(date__gte=start_date)
        if end_date:
            purchase_query = purchase_query.filter(date__lte=end_date)
        
        # Filter by county if specified
        if county:
            purchase_query = purchase_query.filter(county__county=county)
        
        # Get all suppliers
        all_suppliers = Supplier.objects.all()
        if county:
            all_suppliers = all_suppliers.filter(county__county=county)
        
        # Supplier metrics
        supplier_data = []
        total_suppliers = all_suppliers.count()
        active_suppliers = 0
        total_purchase_value = Decimal('0.00')
        total_purchase_orders = 0
        
        for supplier in all_suppliers:
            # Get purchases for this supplier in the period
            supplier_purchases = purchase_query.filter(supplier_id=supplier)
            
            purchase_total = supplier_purchases.aggregate(
                total=Sum('total_amount')
            )['total'] or Decimal('0.00')
            
            orders_count = supplier_purchases.count()
            
            # Calculate payment ratio
            payments_total = supplier.total_payments
            payment_ratio = (payments_total / supplier.total_purchases * 100) if supplier.total_purchases > 0 else 0
            
            # Outstanding balance
            outstanding = supplier.total_purchases - supplier.total_payments
            
            # Get purchase details for this supplier
            supplier_details = PurchaseDetail.objects.filter(
                po_id__in=supplier_purchases
            )
            
            # Calculate delivery reliability (based on shipping status)
            delivered_orders = supplier_purchases.filter(
                shipping_status__shipping_status='DELIVERED'
            ).count()
            delivery_rate = (delivered_orders / orders_count * 100) if orders_count > 0 else 0
            
            # Calculate average lead time (days from order to delivery)
            # This is a simplified calculation
            avg_lead_time = 0
            if orders_count > 0:
                # You can enhance this based on your actual delivery tracking
                avg_lead_time = 7  # Placeholder
            
            # Categories supplied
            categories_supplied = supplier_details.values('item_category').distinct().count()
            
            # Items supplied
            items_supplied = supplier_details.values('item_id').distinct().count()
            
            if purchase_total > 0:
                active_suppliers += 1
                total_purchase_value += purchase_total
                total_purchase_orders += orders_count
                
                # Get first and last order dates
                first_order = supplier_purchases.order_by('date').first()
                last_order = supplier_purchases.order_by('-date').first()
                
                # Calculate average days between orders
                if orders_count > 1 and first_order and last_order:
                    days_span = (last_order.date - first_order.date).days
                    avg_days_between = days_span / (orders_count - 1) if orders_count > 1 else 0
                else:
                    avg_days_between = 0
                
                supplier_data.append({
                    'supplier_id': supplier.supplier_id,
                    'supplier_name': supplier.supplier_name,
                    'phone': supplier.phone_number or '',
                    'email': supplier.email or '',
                    'county': supplier.county.county if supplier.county else '',
                    'town': supplier.town.town if supplier.town else '',
                    'total_purchases': float(purchase_total),
                    'total_orders': orders_count,
                    'avg_order_value': float(purchase_total / orders_count) if orders_count > 0 else 0,
                    'lifetime_value': float(supplier.total_purchases),
                    'total_payments': float(payments_total),
                    'outstanding_balance': float(outstanding),
                    'payment_ratio': float(payment_ratio),
                    'delivery_rate': float(delivery_rate),
                    'avg_lead_time': avg_lead_time,
                    'categories_supplied': categories_supplied,
                    'items_supplied': items_supplied,
                    'first_order_date': first_order.date.strftime('%d/%m/%Y') if first_order else '',
                    'last_order_date': last_order.date.strftime('%d/%m/%Y') if last_order else '',
                    'avg_days_between_orders': int(avg_days_between)
                })
        
        # Sort by total purchases
        supplier_data.sort(key=lambda x: x['total_purchases'], reverse=True)
        
        # Supplier segmentation
        if supplier_data:
            # Top 20% by value = Strategic suppliers
            strategic_threshold_index = max(1, int(len(supplier_data) * 0.2))
            strategic_suppliers = supplier_data[:strategic_threshold_index]
            
            # Next 30% = Preferred suppliers
            preferred_threshold_index = max(strategic_threshold_index + 1, int(len(supplier_data) * 0.5))
            preferred_suppliers = supplier_data[strategic_threshold_index:preferred_threshold_index]
            
            # Remaining = Standard suppliers
            standard_suppliers = supplier_data[preferred_threshold_index:]
            
            segmentation = {
                'strategic': {
                    'count': len(strategic_suppliers),
                    'total_value': sum(s['total_purchases'] for s in strategic_suppliers),
                    'percentage': (len(strategic_suppliers) / len(supplier_data) * 100) if supplier_data else 0
                },
                'preferred': {
                    'count': len(preferred_suppliers),
                    'total_value': sum(s['total_purchases'] for s in preferred_suppliers),
                    'percentage': (len(preferred_suppliers) / len(supplier_data) * 100) if supplier_data else 0
                },
                'standard': {
                    'count': len(standard_suppliers),
                    'total_value': sum(s['total_purchases'] for s in standard_suppliers),
                    'percentage': (len(standard_suppliers) / len(supplier_data) * 100) if supplier_data else 0
                }
            }
        else:
            segmentation = {
                'strategic': {'count': 0, 'total_value': 0, 'percentage': 0},
                'preferred': {'count': 0, 'total_value': 0, 'percentage': 0},
                'standard': {'count': 0, 'total_value': 0, 'percentage': 0}
            }
        
        # Geographic distribution
        geographic_data = all_suppliers.values('county__county').annotate(
            supplier_count=Count('supplier_id'),
            total_purchases=Sum('total_purchases')
        ).order_by('-total_purchases')[:10]
        
        # Performance rating analysis
        performance_analysis = {
            'excellent': 0,  # >95% delivery rate AND >90% payment ratio
            'good': 0,       # >85% delivery rate AND >70% payment ratio
            'fair': 0,       # >70% delivery rate
            'poor': 0        # <70% delivery rate
        }
        
        for supplier in supplier_data:
            delivery = supplier['delivery_rate']
            payment = supplier['payment_ratio']
            
            if delivery >= 95 and payment >= 90:
                performance_analysis['excellent'] += 1
            elif delivery >= 85 and payment >= 70:
                performance_analysis['good'] += 1
            elif delivery >= 70:
                performance_analysis['fair'] += 1
            else:
                performance_analysis['poor'] += 1
        
        # Format response
        report_data = {
            'kpis': {
                'total_suppliers': total_suppliers,
                'active_suppliers': active_suppliers,
                'inactive_suppliers': total_suppliers - active_suppliers,
                'total_purchases': float(total_purchase_value),
                'total_orders': total_purchase_orders,
                'avg_supplier_value': float(total_purchase_value / active_suppliers) if active_suppliers > 0 else 0,
                'avg_orders_per_supplier': float(total_purchase_orders / active_suppliers) if active_suppliers > 0 else 0
            },
            'top_suppliers': supplier_data[:10],  # Top 10 suppliers
            'all_suppliers': supplier_data,  # All supplier data
            'segmentation': segmentation,
            'geographic_distribution': [
                {
                    'county': item['county__county'] or 'Unknown',
                    'supplier_count': item['supplier_count'],
                    'total_purchases': float(item['total_purchases'] or 0)
                }
                for item in geographic_data
            ],
            'performance_rating': performance_analysis,
            'period': {
                'start': start_date.strftime('%d/%m/%Y') if start_date else '',
                'end': end_date.strftime('%d/%m/%Y') if end_date else ''
            }
        }
        
        return JsonResponse({
            'success': True,
            'data': report_data,
            'report_type': 'Supplier Analysis Report'
        })
    
    except Exception as e:
        print(f"Error generating supplier analysis: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)
        
# ============================================
# 6. TAX SUMMARY REPORT
# ============================================

@csrf_exempt
@login_required(login_url='/login/')
def api_generate_tax_summary(request):
    """
    Generate Tax Summary Report
    Returns: Tax collected on sales and paid on purchases
    """
    try:
        start_date_str = request.GET.get('start_date', '')
        end_date_str = request.GET.get('end_date', '')
        
        start_date = parse_date(start_date_str) if start_date_str else None
        end_date = parse_date(end_date_str) if end_date_str else None
        
        # Get Sales Orders in period
        sales_query = SalesOrder.objects.all()
        if start_date:
            sales_query = sales_query.filter(date__gte=start_date)
        if end_date:
            sales_query = sales_query.filter(date__lte=end_date)
        
        # Get Purchase Orders in period
        purchase_query = PurchaseOrder.objects.all()
        if start_date:
            purchase_query = purchase_query.filter(date__gte=start_date)
        if end_date:
            purchase_query = purchase_query.filter(date__lte=end_date)
        
        # Calculate Sales Tax Collected
        sales_details = SalesDetail.objects.filter(so_id__in=sales_query)
        
        total_sales_tax = Decimal('0.00')
        sales_by_rate = {}
        sales_transactions = []
        
        for detail in sales_details:
            # Calculate tax for this detail
            subtotal = detail.quantity_sold * detail.unit_price
            tax_amount = subtotal * detail.tax_rate / 100
            total_sales_tax += tax_amount
            
            # Group by tax rate
            rate_key = float(detail.tax_rate)
            if rate_key not in sales_by_rate:
                sales_by_rate[rate_key] = {
                    'tax_rate': rate_key,
                    'taxable_amount': Decimal('0.00'),
                    'tax_collected': Decimal('0.00'),
                    'transactions': 0
                }
            
            sales_by_rate[rate_key]['taxable_amount'] += subtotal
            sales_by_rate[rate_key]['tax_collected'] += tax_amount
            sales_by_rate[rate_key]['transactions'] += 1
            
            # Add to transactions list
            sales_transactions.append({
                'date': detail.date.strftime('%d/%m/%Y') if detail.date else '',
                'invoice': detail.invoice_number,
                'customer': detail.customer_name,
                'item': detail.item_name,
                'amount': float(subtotal),
                'tax_rate': float(detail.tax_rate),
                'tax': float(tax_amount)
            })
        
        # Calculate Purchase Tax Paid
        purchase_details = PurchaseDetail.objects.filter(po_id__in=purchase_query)
        
        total_purchase_tax = Decimal('0.00')
        purchase_by_rate = {}
        purchase_transactions = []
        
        for detail in purchase_details:
            # Calculate tax for this detail
            subtotal = detail.quantity_purchased * detail.unit_cost
            tax_amount = subtotal * detail.tax_rate / 100
            total_purchase_tax += tax_amount
            
            # Group by tax rate
            rate_key = float(detail.tax_rate)
            if rate_key not in purchase_by_rate:
                purchase_by_rate[rate_key] = {
                    'tax_rate': rate_key,
                    'taxable_amount': Decimal('0.00'),
                    'tax_paid': Decimal('0.00'),
                    'transactions': 0
                }
            
            purchase_by_rate[rate_key]['taxable_amount'] += subtotal
            purchase_by_rate[rate_key]['tax_paid'] += tax_amount
            purchase_by_rate[rate_key]['transactions'] += 1
            
            # Add to transactions list
            purchase_transactions.append({
                'date': detail.date.strftime('%d/%m/%Y') if detail.date else '',
                'bill': detail.bill_number,
                'supplier': detail.supplier_name,
                'item': detail.item_name,
                'amount': float(subtotal),
                'tax_rate': float(detail.tax_rate),
                'tax': float(tax_amount)
            })
        
        # Net Tax Position (Tax Collected - Tax Paid)
        net_tax = total_sales_tax - total_purchase_tax
        
        # Convert to list for JSON
        sales_by_rate_list = [
            {
                'tax_rate': data['tax_rate'],
                'taxable_amount': float(data['taxable_amount']),
                'tax_collected': float(data['tax_collected']),
                'transactions': data['transactions']
            }
            for data in sorted(sales_by_rate.values(), key=lambda x: x['tax_rate'])
        ]
        
        purchase_by_rate_list = [
            {
                'tax_rate': data['tax_rate'],
                'taxable_amount': float(data['taxable_amount']),
                'tax_paid': float(data['tax_paid']),
                'transactions': data['transactions']
            }
            for data in sorted(purchase_by_rate.values(), key=lambda x: x['tax_rate'])
        ]
        
        report_data = {
            'kpis': {
                'total_sales_tax': float(total_sales_tax),
                'total_purchase_tax': float(total_purchase_tax),
                'net_tax': float(net_tax),
                'sales_transactions': len(sales_transactions),
                'purchase_transactions': len(purchase_transactions)
            },
            'sales_by_rate': sales_by_rate_list,
            'purchase_by_rate': purchase_by_rate_list,
            'sales_transactions': sorted(sales_transactions, key=lambda x: x['date'], reverse=True)[:50],
            'purchase_transactions': sorted(purchase_transactions, key=lambda x: x['date'], reverse=True)[:50],
            'period': {
                'start': start_date.strftime('%d/%m/%Y') if start_date else '',
                'end': end_date.strftime('%d/%m/%Y') if end_date else ''
            }
        }
        
        return JsonResponse({
            'success': True,
            'data': report_data,
            'report_type': 'Tax Summary Report'
        })
    
    except Exception as e:
        print(f"❌ Error generating tax summary: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)
        
# ============================================
# EXPORT TO PDF
# ============================================

@csrf_exempt
@login_required(login_url='/login/')
def api_export_report_pdf(request):
    """
    Export any report to PDF
    Accepts: report_type, report_data (JSON)
    Returns: PDF file download
    """
    try:
        if request.method != 'POST':
            return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
        
        data = json.loads(request.body)
        report_type = data.get('report_type', 'Report')
        report_data = data.get('report_data', {})
        period = data.get('period', {})
        
        # Create PDF buffer
        from io import BytesIO
        buffer = BytesIO()
        
        # Create PDF document
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=30,
            leftMargin=30,
            topMargin=50,
            bottomMargin=50
        )
        
        # Container for PDF elements
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=28,
            textColor=colors.HexColor('#2c3e50'),
            spaceAfter=12,
            alignment=TA_CENTER,
            underlineColor=colors.HexColor('#2c3e50')
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#1abc9c'),
            spaceAfter=10,
            spaceBefore=20
        )
        
        # Title
        title = Paragraph(report_type, title_style)
        elements.append(title)
        elements.append(Spacer(1, 12))
        
        # Period
        if period.get('start') and period.get('end'):
            period_text = f"Period: {period['start']} - {period['end']}"
            period_para = Paragraph(period_text, styles['Normal'])
            elements.append(period_para)
            elements.append(Spacer(1, 20))
        
        # KPIs Section
        if 'kpis' in report_data:
            elements.append(Paragraph('Key Performance Indicators', heading_style))
            
            kpi_data = []
            for key, value in report_data['kpis'].items():
                label = key.replace('_', ' ').title()
                if isinstance(value, (int, float)):
                    formatted_value = f"KSH {value:,.2f}" if 'total' in key or 'amount' in key or 'balance' in key else str(value)
                else:
                    formatted_value = str(value)
                kpi_data.append([label, formatted_value])
            
            kpi_table = Table(kpi_data, colWidths=[3*inch, 2*inch])
            kpi_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f8f9fa')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#2c3e50')),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#ecf0f1'))
            ]))
            elements.append(kpi_table)
            elements.append(Spacer(1, 20))
        
        # Add specific report sections based on report_type
        if 'sales' in report_type.lower():
            _add_sales_sections(elements, report_data, heading_style, styles)
        elif 'inventory' in report_type.lower():
            _add_inventory_sections(elements, report_data, heading_style, styles)
        elif 'profit' in report_type.lower():
            _add_profit_loss_sections(elements, report_data, heading_style, styles)
        elif 'purchase' in report_type.lower():
            _add_purchase_sections(elements, report_data, heading_style, styles)
        elif 'outstanding' in report_type.lower():
            _add_outstanding_sections(elements, report_data, heading_style, styles)
        elif 'tax' in report_type.lower():
            _add_tax_sections(elements, report_data, heading_style, styles)
        
        # Build PDF
        doc.build(elements)
        
        # Return PDF as download
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        filename = f"{report_type.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    except Exception as e:
        print(f"❌ Error exporting to PDF: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


# Helper functions for PDF sections
def _add_sales_sections(elements, report_data, heading_style, styles):
    """Add sales-specific sections to PDF"""
    # Top Customers
    if 'top_customers' in report_data:
        elements.append(Paragraph('Top Customers', heading_style))
        
        customer_data = [['Customer Name', 'Total Sales', 'Orders', 'Avg Order']]
        for customer in report_data['top_customers'][:10]:
            customer_data.append([
                customer['name'],
                f"KSH {customer['total']:,.2f}",
                str(customer['orders']),
                f"KSH {(customer['total'] / customer['orders']):,.2f}"
            ])
        
        table = Table(customer_data, colWidths=[2*inch, 1.5*inch, 1*inch, 1.5*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey)
        ]))
        elements.append(table)


def _add_inventory_sections(elements, report_data, heading_style, styles):
    """Add inventory-specific sections to PDF"""
    # Low Stock Items
    if 'low_stock' in report_data and report_data['low_stock']:
        elements.append(Paragraph('⚠️ Items Requiring Reorder', heading_style))
        
        stock_data = [['Item ID', 'Item Name', 'Current Stock', 'Reorder Level']]
        for item in report_data['low_stock'][:15]:
            stock_data.append([
                item['item_id'],
                item['name'],
                str(item['remaining_qty']),
                str(item['reorder_level'])
            ])
        
        table = Table(stock_data, colWidths=[1.5*inch, 2.5*inch, 1.5*inch, 1.5*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e74c3c')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (2, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey)
        ]))
        elements.append(table)


def _add_profit_loss_sections(elements, report_data, heading_style, styles):
    """Add P&L-specific sections to PDF"""
    if 'breakdown' in report_data:
        elements.append(Paragraph('Income Statement', heading_style))
        
        pl_data = [
            ['Revenue', f"KSH {report_data['breakdown']['revenue']:,.2f}"],
            ['Less: Cost of Goods Sold', f"(KSH {report_data['breakdown']['cogs']:,.2f})"],
            ['Gross Profit', f"KSH {report_data['breakdown']['gross_profit']:,.2f}"],
            ['Less: Operating Expenses', f"(KSH {report_data['breakdown']['shipping_expense']:,.2f})"],
            ['Net Profit', f"KSH {report_data['breakdown']['net_profit']:,.2f}"]
        ]
        
        table = Table(pl_data, colWidths=[3*inch, 2*inch])
        table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('LINEABOVE', (0, -1), (-1, -1), 2, colors.HexColor('#2c3e50')),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#d1f2eb'))
        ]))
        elements.append(table)


def _add_purchase_sections(elements, report_data, heading_style, styles):
    """Add purchase-specific sections to PDF"""
    if 'top_suppliers' in report_data:
        elements.append(Paragraph('Top Suppliers', heading_style))
        
        supplier_data = [['Supplier Name', 'Total Purchases', 'Orders', 'Avg Order']]
        for supplier in report_data['top_suppliers'][:10]:
            supplier_data.append([
                supplier['name'],
                f"KSH {supplier['total']:,.2f}",
                str(supplier['orders']),
                f"KSH {(supplier['total'] / supplier['orders']):,.2f}"
            ])
        
        table = Table(supplier_data, colWidths=[2*inch, 1.5*inch, 1*inch, 1.5*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey)
        ]))
        elements.append(table)


def _add_outstanding_sections(elements, report_data, heading_style, styles):
    """Add outstanding balances sections to PDF"""
    if 'customer_balances' in report_data and report_data['customer_balances']:
        elements.append(Paragraph('Customer Balances (Receivable)', heading_style))
        
        customer_data = [['Customer ID', 'Customer Name', 'Total Sales', 'Payments', 'Balance']]
        for customer in report_data['customer_balances'][:15]:
            customer_data.append([
                customer['id'],
                customer['name'],
                f"KSH {customer['total_sales']:,.2f}",
                f"KSH {customer['total_payments']:,.2f}",
                f"KSH {customer['balance']:,.2f}"
            ])
        
        table = Table(customer_data, colWidths=[1*inch, 1.8*inch, 1.4*inch, 1.4*inch, 1.4*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2ecc71')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey)
        ]))
        elements.append(table)
        elements.append(Spacer(1, 20))
    
    if 'supplier_balances' in report_data and report_data['supplier_balances']:
        elements.append(Paragraph('Supplier Balances (Payable)', heading_style))
        
        supplier_data = [['Supplier ID', 'Supplier Name', 'Total Purchases', 'Payments', 'Balance']]
        for supplier in report_data['supplier_balances'][:15]:
            supplier_data.append([
                supplier['id'],
                supplier['name'],
                f"KSH {supplier['total_purchases']:,.2f}",
                f"KSH {supplier['total_payments']:,.2f}",
                f"KSH {supplier['balance']:,.2f}"
            ])
        
        table = Table(supplier_data, colWidths=[1*inch, 1.8*inch, 1.4*inch, 1.4*inch, 1.4*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f39c12')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey)
        ]))
        elements.append(table)


def _add_tax_sections(elements, report_data, heading_style, styles):
    """Add tax summary sections to PDF"""
    if 'sales_by_rate' in report_data:
        elements.append(Paragraph('Sales Tax by Rate', heading_style))
        
        tax_data = [['Tax Rate', 'Taxable Amount', 'Tax Collected', 'Transactions']]
        for rate in report_data['sales_by_rate']:
            tax_data.append([
                f"{rate['tax_rate']}%",
                f"KSH {rate['taxable_amount']:,.2f}",
                f"KSH {rate['tax_collected']:,.2f}",
                str(rate['transactions'])
            ])
        
        table = Table(tax_data, colWidths=[1.5*inch, 2*inch, 2*inch, 1.5*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey)
        ]))
        elements.append(table)
        
# ============================================
# EXPORT TO EXCEL
# ============================================
@csrf_exempt
@login_required(login_url='/login/')
def api_export_report_excel(request):
    """
    Export any report to Excel
    Accepts: report_type, report_data (JSON)
    Returns: Excel file download
    """
    try:
        if request.method != 'POST':
            return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
        
        data = json.loads(request.body)
        report_type = data.get('report_type', 'Report')
        report_data = data.get('report_data', {})
        period = data.get('period', {})
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"
        
        # Styles
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=16, color="1ABC9C")
        
        # Title
        ws['A1'] = report_type
        ws['A1'].font = title_font
        ws.merge_cells('A1:E1')
        
        # Period
        if period.get('start') and period.get('end'):
            ws['A2'] = f"Period: {period['start']} - {period['end']}"
            ws.merge_cells('A2:E2')
        
        current_row = 4
        
        # KPIs Section
        if 'kpis' in report_data:
            ws[f'A{current_row}'] = 'KEY PERFORMANCE INDICATORS'
            ws[f'A{current_row}'].font = Font(bold=True, size=12)
            current_row += 1
            
            for key, value in report_data['kpis'].items():
                label = key.replace('_', ' ').title()
                ws[f'A{current_row}'] = label
                if isinstance(value, (int, float)):
                    ws[f'B{current_row}'] = value
                    ws[f'B{current_row}'].number_format = '#,##0.00'
                else:
                    ws[f'B{current_row}'] = str(value)
                current_row += 1
            
            current_row += 2
        
        # Add specific sections based on report type
        report_type_lower = report_type.lower()
        
        if 'sales' in report_type_lower and 'summary' in report_type_lower:
            current_row = _add_excel_sales_sections(ws, report_data, current_row, header_fill, header_font)
        elif 'inventory' in report_type_lower:
            current_row = _add_excel_inventory_sections(ws, report_data, current_row, header_fill, header_font)
        elif 'profit' in report_type_lower or 'loss' in report_type_lower:
            current_row = _add_excel_profit_loss_sections(ws, report_data, current_row, header_fill, header_font)
        elif 'purchase' in report_type_lower:
            current_row = _add_excel_purchase_sections(ws, report_data, current_row, header_fill, header_font)
        elif 'outstanding' in report_type_lower:
            current_row = _add_excel_outstanding_sections(ws, report_data, current_row, header_fill, header_font)
        elif 'customer' in report_type_lower and 'analysis' in report_type_lower:
            current_row = _add_excel_customer_analysis_sections(ws, report_data, current_row, header_fill, header_font)
        elif 'supplier' in report_type_lower and 'analysis' in report_type_lower:
            current_row = _add_excel_supplier_analysis_sections(ws, report_data, current_row, header_fill, header_font)
        elif 'tax' in report_type_lower:
            current_row = _add_excel_tax_sections(ws, report_data, current_row, header_fill, header_font)
        
        # Adjust column widths (CORRECTED INDENTATION)
        from openpyxl.cell.cell import MergedCell
        
        for col in ws.columns:
            max_length = 0
            column_letter = None
            
            for cell in col:
                # Skip merged cells
                if isinstance(cell, MergedCell):
                    continue
                
                # Get column letter from first non-merged cell
                if column_letter is None:
                    column_letter = cell.column_letter
                
                # Calculate max length
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # Set column width if we found a valid column letter
            if column_letter:
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Return as download
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"{report_type.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    except Exception as e:
        print(f"❌ Error exporting to Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
# Helper functions for Excel sections
def _add_excel_sales_sections(ws, report_data, start_row, header_fill, header_font):
    """Add sales data to Excel"""
    if 'top_customers' in report_data:
        ws[f'A{start_row}'] = 'TOP CUSTOMERS'
        ws[f'A{start_row}'].font = Font(bold=True, size=12)
        start_row += 1
        
        headers = ['Customer Name', 'Total Sales', 'Orders', 'Avg Order Value']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        start_row += 1
        for customer in report_data['top_customers']:
            ws.cell(row=start_row, column=1, value=customer['name'])
            ws.cell(row=start_row, column=2, value=customer['total']).number_format = '#,##0.00'
            ws.cell(row=start_row, column=3, value=customer['orders'])
            ws.cell(row=start_row, column=4, value=customer['total'] / customer['orders']).number_format = '#,##0.00'
            start_row += 1
    
    return start_row

def _add_excel_inventory_sections(ws, report_data, start_row, header_fill, header_font):
    """Add inventory data to Excel"""
    if 'low_stock' in report_data and report_data['low_stock']:
        ws[f'A{start_row}'] = 'ITEMS REQUIRING REORDER'
        ws[f'A{start_row}'].font = Font(bold=True, size=12, color="E74C3C")
        start_row += 1
        
        headers = ['Item ID', 'Item Name', 'Category', 'Current Stock', 'Reorder Level']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        start_row += 1
        for item in report_data['low_stock']:
            ws.cell(row=start_row, column=1, value=item['item_id'])
            ws.cell(row=start_row, column=2, value=item['name'])
            ws.cell(row=start_row, column=3, value=item['category'])
            ws.cell(row=start_row, column=4, value=item['remaining_qty'])
            ws.cell(row=start_row, column=5, value=item['reorder_level'])
            start_row += 1
    
    return start_row

def _add_excel_profit_loss_sections(ws, report_data, start_row, header_fill, header_font):
    """Add P&L data to Excel"""
    if 'breakdown' in report_data:
        ws[f'A{start_row}'] = 'INCOME STATEMENT'
        ws[f'A{start_row}'].font = Font(bold=True, size=12)
        start_row += 1
        
        pl_items = [
            ('Revenue', report_data['breakdown']['revenue']),
            ('Cost of Goods Sold', -report_data['breakdown']['cogs']),
            ('Gross Profit', report_data['breakdown']['gross_profit']),
            ('Operating Expenses', -report_data['breakdown']['shipping_expense']),
            ('Net Profit', report_data['breakdown']['net_profit'])
        ]
        
        for label, value in pl_items:
            ws.cell(row=start_row, column=1, value=label)
            ws.cell(row=start_row, column=2, value=value).number_format = '#,##0.00'
            start_row += 1
    
    return start_row

def _add_excel_purchase_sections(ws, report_data, start_row, header_fill, header_font):
    """Add purchase data to Excel"""
    if 'top_suppliers' in report_data:
        ws[f'A{start_row}'] = 'TOP SUPPLIERS'
        ws[f'A{start_row}'].font = Font(bold=True, size=12)
        start_row += 1
        
        headers = ['Supplier Name', 'Total Purchases', 'Orders', 'Avg Order Value']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        start_row += 1
        for supplier in report_data['top_suppliers']:
            ws.cell(row=start_row, column=1, value=supplier['name'])
            ws.cell(row=start_row, column=2, value=supplier['total']).number_format = '#,##0.00'
            ws.cell(row=start_row, column=3, value=supplier['orders'])
            ws.cell(row=start_row, column=4, value=supplier['total'] / supplier['orders']).number_format = '#,##0.00'
            start_row += 1
    
    return start_row

def _add_excel_outstanding_sections(ws, report_data, start_row, header_fill, header_font):
    """Add outstanding balances to Excel"""
    if 'customer_balances' in report_data and report_data['customer_balances']:
        ws[f'A{start_row}'] = 'CUSTOMER BALANCES (RECEIVABLE)'
        ws[f'A{start_row}'].font = Font(bold=True, size=12)
        start_row += 1
        
        headers = ['Customer ID', 'Customer Name', 'Total Sales', 'Payments', 'Balance']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        start_row += 1
        for customer in report_data['customer_balances']:
            ws.cell(row=start_row, column=1, value=customer['id'])
            ws.cell(row=start_row, column=2, value=customer['name'])
            ws.cell(row=start_row, column=3, value=customer['total_sales']).number_format = '#,##0.00'
            ws.cell(row=start_row, column=4, value=customer['total_payments']).number_format = '#,##0.00'
            ws.cell(row=start_row, column=5, value=customer['balance']).number_format = '#,##0.00'
            start_row += 1
        
        start_row += 2
    
    if 'supplier_balances' in report_data and report_data['supplier_balances']:
        ws[f'A{start_row}'] = 'SUPPLIER BALANCES (PAYABLE)'
        ws[f'A{start_row}'].font = Font(bold=True, size=12)
        start_row += 1
        
        headers = ['Supplier ID', 'Supplier Name', 'Total Purchases', 'Payments', 'Balance']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        start_row += 1
        for supplier in report_data['supplier_balances']:
            ws.cell(row=start_row, column=1, value=supplier['id'])
            ws.cell(row=start_row, column=2, value=supplier['name'])
            ws.cell(row=start_row, column=3, value=supplier['total_purchases']).number_format = '#,##0.00'
            ws.cell(row=start_row, column=4, value=supplier['total_payments']).number_format = '#,##0.00'
            ws.cell(row=start_row, column=5, value=supplier['balance']).number_format = '#,##0.00'
            start_row += 1
    
    return start_row

def _add_excel_tax_sections(ws, report_data, start_row, header_fill, header_font):
    """Add tax summary to Excel"""
    if 'sales_by_rate' in report_data:
        ws[f'A{start_row}'] = 'SALES TAX BY RATE'
        ws[f'A{start_row}'].font = Font(bold=True, size=12)
        start_row += 1
        
        headers = ['Tax Rate', 'Taxable Amount', 'Tax Collected', 'Transactions']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        start_row += 1
        for rate in report_data['sales_by_rate']:
            ws.cell(row=start_row, column=1, value=f"{rate['tax_rate']}%")
            ws.cell(row=start_row, column=2, value=rate['taxable_amount']).number_format = '#,##0.00'
            ws.cell(row=start_row, column=3, value=rate['tax_collected']).number_format = '#,##0.00'
            ws.cell(row=start_row, column=4, value=rate['transactions'])
            start_row += 1
        
        start_row += 2
    
    if 'purchase_by_rate' in report_data:
        ws[f'A{start_row}'] = 'PURCHASE TAX BY RATE'
        ws[f'A{start_row}'].font = Font(bold=True, size=12)
        start_row += 1
        
        headers = ['Tax Rate', 'Taxable Amount', 'Tax Paid', 'Transactions']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        start_row += 1
        for rate in report_data['purchase_by_rate']:
            ws.cell(row=start_row, column=1, value=f"{rate['tax_rate']}%")
            ws.cell(row=start_row, column=2, value=rate['taxable_amount']).number_format = '#,##0.00'
            ws.cell(row=start_row, column=3, value=rate['tax_paid']).number_format = '#,##0.00'
            ws.cell(row=start_row, column=4, value=rate['transactions'])
            start_row += 1
        
        start_row += 2
    
    if 'kpis' in report_data:
        ws[f'A{start_row}'] = 'NET TAX POSITION'
        ws[f'A{start_row}'].font = Font(bold=True, size=12)
        start_row += 1
        
        ws.cell(row=start_row, column=1, value='Tax Collected (Sales)')
        ws.cell(row=start_row, column=2, value=report_data['kpis']['total_sales_tax']).number_format = '#,##0.00'
        start_row += 1
        
        ws.cell(row=start_row, column=1, value='Tax Paid (Purchases)')
        ws.cell(row=start_row, column=2, value=-report_data['kpis']['total_purchase_tax']).number_format = '#,##0.00'
        start_row += 1
        
        ws.cell(row=start_row, column=1, value='Net Tax Position').font = Font(bold=True)
        ws.cell(row=start_row, column=2, value=report_data['kpis']['net_tax']).number_format = '#,##0.00'
        ws.cell(row=start_row, column=2).font = Font(bold=True)
        start_row += 1
    
    return start_row

def _add_excel_customer_analysis_sections(ws, report_data, start_row, header_fill, header_font):
    """Add customer analysis data to Excel"""
    if 'top_customers' in report_data:
        ws[f'A{start_row}'] = 'TOP CUSTOMERS'
        ws[f'A{start_row}'].font = Font(bold=True, size=12)
        start_row += 1
        
        headers = ['Customer Name', 'Total Sales', 'Orders', 'Avg Order', 'Outstanding', 'Payment Ratio']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        start_row += 1
        for customer in report_data['top_customers']:
            ws.cell(row=start_row, column=1, value=customer['customer_name'])
            ws.cell(row=start_row, column=2, value=customer['total_sales']).number_format = '#,##0.00'
            ws.cell(row=start_row, column=3, value=customer['total_orders'])
            ws.cell(row=start_row, column=4, value=customer['avg_order_value']).number_format = '#,##0.00'
            ws.cell(row=start_row, column=5, value=customer['outstanding_balance']).number_format = '#,##0.00'
            ws.cell(row=start_row, column=6, value=customer['payment_ratio']).number_format = '0.0"%"'
            start_row += 1
    
    return start_row

def _add_excel_supplier_analysis_sections(ws, report_data, start_row, header_fill, header_font):
    """Add supplier analysis data to Excel"""
    if 'top_suppliers' in report_data:
        ws[f'A{start_row}'] = 'TOP SUPPLIERS'
        ws[f'A{start_row}'].font = Font(bold=True, size=12)
        start_row += 1
        
        headers = ['Supplier Name', 'Total Purchases', 'Orders', 'Avg Order', 'Outstanding', 'Delivery Rate']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        start_row += 1
        for supplier in report_data['top_suppliers']:
            ws.cell(row=start_row, column=1, value=supplier['supplier_name'])
            ws.cell(row=start_row, column=2, value=supplier['total_purchases']).number_format = '#,##0.00'
            ws.cell(row=start_row, column=3, value=supplier['total_orders'])
            ws.cell(row=start_row, column=4, value=supplier['avg_order_value']).number_format = '#,##0.00'
            ws.cell(row=start_row, column=5, value=supplier['outstanding_balance']).number_format = '#,##0.00'
            ws.cell(row=start_row, column=6, value=supplier['delivery_rate']).number_format = '0.0"%"'
            start_row += 1
    
    return start_row

# ====================================== SETTINGS MODULE VIEWS ==================================================================================================

@csrf_exempt
@login_required(login_url='/login/')
def api_get_user_profile(request):
    """Get current user profile"""
    try:
        user = request.user
        
        profile_data = {
            'full_name': user.full_name,
            'email': user.email,
            'phone_number': user.phone_number or '',
            'user_role': user.user_role.user_role if user.user_role else 'User',
            'is_admin': user.is_admin
        }
        
        return JsonResponse({
            'success': True,
            'data': profile_data
        })
    
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
@login_required(login_url='/login/')
def api_update_user_profile(request):
    """Update user profile"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
    
    try:
        data = json.loads(request.body)
        user = request.user
        
        # Update fields
        user.full_name = data.get('full_name', user.full_name)
        user.phone_number = data.get('phone_number', user.phone_number)
        user.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Profile updated successfully'
        })
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@csrf_exempt
@login_required(login_url='/login/')
def api_change_password(request):
    """Change user password"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
    
    try:
        data = json.loads(request.body)
        user = request.user
        
        current_password = data.get('current_password', '')
        new_password = data.get('new_password', '')
        
        # Verify current password
        if not user.check_password(current_password):
            return JsonResponse({
                'success': False,
                'message': 'Current password is incorrect'
            }, status=400)
        
        # Validate new password
        if len(new_password) < 6:
            return JsonResponse({
                'success': False,
                'message': 'New password must be at least 6 characters'
            }, status=400)
        
        # Update password
        user.set_password(new_password)
        user.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Password changed successfully'
        })
    
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)